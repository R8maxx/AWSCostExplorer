#!/usr/bin/env python3
"""
Script para extraer costos de AWS por etiqueta Name
Incluye todos los servicios (EC2, S3, RDS, etc.) y AWS Backup
Exporta los resultados a Excel
"""

import boto3
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
import argparse
import sys


def obtener_rango_fechas(mes=None, anio=None):
    """
    Obtiene el rango de fechas para la consulta
    Si no se especifica mes/año, usa el mes actual
    """
    if mes and anio:
        fecha_inicio = datetime(anio, mes, 1)
    else:
        ahora = datetime.now()
        fecha_inicio = datetime(ahora.year, ahora.month, 1)

    # Calcular el último día del mes
    if fecha_inicio.month == 12:
        fecha_fin = datetime(fecha_inicio.year + 1, 1, 1)
    else:
        fecha_fin = datetime(fecha_inicio.year, fecha_inicio.month + 1, 1)

    return fecha_inicio.strftime('%Y-%m-%d'), fecha_fin.strftime('%Y-%m-%d')


def obtener_costos_por_servicio(cliente_ce, fecha_inicio, fecha_fin):
    """
    Obtiene los costos agrupados por servicio y etiquetas.
    AWS solo permite 2 GroupBy por consulta, así que hacemos múltiples consultas.
    """
    print(f"Obteniendo costos desde {fecha_inicio} hasta {fecha_fin}...")
    print("Esto puede tardar unos momentos...")

    try:
        # Primera consulta: Servicio + Name
        print("  → Consultando costos por Name...")
        response_name = cliente_ce.get_cost_and_usage(
            TimePeriod={
                'Start': fecha_inicio,
                'End': fecha_fin
            },
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'DIMENSION', 'Key': 'SERVICE'},
                {'Type': 'TAG', 'Key': 'Name'}
            ]
        )

        # Crear un diccionario para almacenar costos por Name
        costos_por_name = {}
        for periodo in response_name['ResultsByTime']:
            for grupo in periodo['Groups']:
                servicio = grupo['Keys'][0]
                name = grupo['Keys'][1].replace('Name$', '') if len(grupo['Keys']) > 1 else 'Sin etiqueta'
                if name == 'Name$':
                    name = 'Sin etiqueta'
                costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                if costo > 0:
                    if name not in costos_por_name:
                        costos_por_name[name] = {}
                    costos_por_name[name][servicio] = costos_por_name[name].get(servicio, 0) + costo

        # Segunda consulta: ServerGroup
        print("  → Consultando etiquetas ServerGroup...")
        etiquetas_recursos = {}

        try:
            response_sg = cliente_ce.get_cost_and_usage(
                TimePeriod={
                    'Start': fecha_inicio,
                    'End': fecha_fin
                },
                Granularity='MONTHLY',
                Metrics=['UnblendedCost'],
                GroupBy=[
                    {'Type': 'TAG', 'Key': 'Name'},
                    {'Type': 'TAG', 'Key': 'ServerGroup'}
                ]
            )

            for periodo in response_sg['ResultsByTime']:
                for grupo in periodo['Groups']:
                    name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][0] != 'Name$' else 'Sin etiqueta'
                    servergroup = grupo['Keys'][1].replace('ServerGroup$', '') if len(grupo['Keys']) > 1 and \
                                                                                  grupo['Keys'][
                                                                                      1] != 'ServerGroup$' else ''

                    if name and name != 'Sin etiqueta':
                        if name not in etiquetas_recursos:
                            etiquetas_recursos[name] = {'servergroup': ''}
                        if servergroup:
                            etiquetas_recursos[name]['servergroup'] = servergroup
        except Exception as e:
            print(f"    ⚠ Advertencia: No se pudo obtener ServerGroup: {e}")

        return costos_por_name, etiquetas_recursos

    except Exception as e:
        print(f"Error al obtener costos: {e}")
        sys.exit(1)


def obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin):
    """
    Obtiene los costos específicos de AWS Backup con sus etiquetas
    """
    print("Obteniendo costos de AWS Backup...")

    costos_backup = {}

    # Mapeo de planes de backup
    planes_backup = {
        'BackupDia': 'avanza_backup_daily',
        'BackupSemana': 'avanza_backup_weekly',
        'BackupMes': 'avanza-backup-monthly'
    }

    try:
        # Obtener costos de AWS Backup filtrados por etiqueta AWSBackup
        for valor_etiqueta, nombre_plan in planes_backup.items():
            response = cliente_ce.get_cost_and_usage(
                TimePeriod={
                    'Start': fecha_inicio,
                    'End': fecha_fin
                },
                Granularity='MONTHLY',
                Metrics=['UnblendedCost'],
                Filter={
                    'And': [
                        {
                            'Dimensions': {
                                'Key': 'SERVICE',
                                'Values': ['AWS Backup']
                            }
                        },
                        {
                            'Tags': {
                                'Key': 'AWSBackup',
                                'Values': [valor_etiqueta]
                            }
                        }
                    ]
                },
                GroupBy=[
                    {'Type': 'TAG', 'Key': 'Name'}
                ]
            )

            for periodo in response['ResultsByTime']:
                for grupo in periodo['Groups']:
                    etiqueta_name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][
                                                                                 0] != 'Name$' else 'Sin etiqueta'
                    costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                    if costo > 0:
                        clave = f"{etiqueta_name}|{nombre_plan}"
                        costos_backup[clave] = costos_backup.get(clave, 0) + costo

        return costos_backup
    except Exception as e:
        print(f"Advertencia: No se pudieron obtener costos de AWS Backup: {e}")
        return {}


def obtener_desglose_ec2(cliente_ce, fecha_inicio, fecha_fin):
    """
    Obtiene el desglose detallado de costos de EC2 por tipo de uso (Usage Type)
    Incluye todos los servicios relacionados con EC2
    """
    print("  → Obteniendo desglose detallado de EC2...")

    desglose_ec2 = defaultdict(lambda: defaultdict(float))

    # Lista de servicios EC2 a desglosar
    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store',
        'Amazon EC2 Container Registry (ECR)'
    ]

    try:
        for servicio_ec2 in servicios_ec2:
            try:
                # Consultar cada servicio EC2 con desglose por Usage Type y Name
                response = cliente_ce.get_cost_and_usage(
                    TimePeriod={
                        'Start': fecha_inicio,
                        'End': fecha_fin
                    },
                    Granularity='MONTHLY',
                    Metrics=['UnblendedCost'],
                    Filter={
                        'Dimensions': {
                            'Key': 'SERVICE',
                            'Values': [servicio_ec2]
                        }
                    },
                    GroupBy=[
                        {'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'},
                        {'Type': 'TAG', 'Key': 'Name'}
                    ]
                )

                for periodo in response['ResultsByTime']:
                    for grupo in periodo['Groups']:
                        usage_type = grupo['Keys'][0]
                        name = grupo['Keys'][1].replace('Name$', '') if len(grupo['Keys']) > 1 else 'Sin etiqueta'
                        if name == 'Name$':
                            name = 'Sin etiqueta'
                        costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                        if costo > 0:
                            # Categorizar el tipo de uso
                            servicio_detallado = categorizar_ec2_usage(usage_type)
                            desglose_ec2[name][servicio_detallado] += costo

            except Exception as e:
                print(f"    ⚠ No se pudo obtener desglose de {servicio_ec2}: {e}")
                continue

        return desglose_ec2
    except Exception as e:
        print(f"    ⚠ Advertencia: No se pudo obtener desglose de EC2: {e}")
        return {}


def categorizar_ec2_usage(usage_type):
    """
    Categoriza los tipos de uso de EC2 en categorías legibles
    """
    usage_type_lower = usage_type.lower()

    # Instancias EC2
    if 'boxusage' in usage_type_lower or 'instanceusage' in usage_type_lower or 'hoursusage' in usage_type_lower:
        return 'EC2 - Instancias'

    # EBS Volumes (primero porque puede tener snapshot también)
    elif 'volumeusage' in usage_type_lower or 'volumepiops' in usage_type_lower:
        return 'EC2 - EBS Volumes'

    # Snapshots
    elif 'snapshot' in usage_type_lower or 'ebssnapshot' in usage_type_lower:
        return 'EC2 - EBS Snapshots'

    # Elastic IPs
    elif 'elasticip' in usage_type_lower or 'idleaddress' in usage_type_lower or 'addressusage' in usage_type_lower:
        return 'EC2 - Elastic IPs'

    # Network Interfaces (ENI)
    elif 'networkinterface' in usage_type_lower or 'vpcendpoint' in usage_type_lower:
        return 'EC2 - Network Interfaces'

    # Data Transfer
    elif 'datatransfer' in usage_type_lower or 'data-transfer' in usage_type_lower or 'regional-bytes' in usage_type_lower:
        if 'in' in usage_type_lower and 'out' not in usage_type_lower:
            return 'EC2 - Data Transfer In'
        elif 'out' in usage_type_lower:
            return 'EC2 - Data Transfer Out'
        elif 'regional' in usage_type_lower or 'inter' in usage_type_lower:
            return 'EC2 - Data Transfer Regional'
        else:
            return 'EC2 - Data Transfer'

    # NAT Gateway
    elif 'natgateway' in usage_type_lower:
        if 'bytes' in usage_type_lower:
            return 'EC2 - NAT Gateway (Data)'
        else:
            return 'EC2 - NAT Gateway (Hours)'

    # Load Balancers
    elif 'loadbalancer' in usage_type_lower or 'elb' in usage_type_lower or 'lcu' in usage_type_lower:
        return 'EC2 - Load Balancer'

    # VPN
    elif 'vpn' in usage_type_lower:
        return 'EC2 - VPN Connection'

    # IOPS provisionadas
    elif 'iops' in usage_type_lower or 'piops' in usage_type_lower:
        return 'EC2 - EBS IOPS Provisionadas'

    # Throughput
    elif 'throughput' in usage_type_lower or 'volumethroughput' in usage_type_lower:
        return 'EC2 - EBS Throughput'

    # EBS Optimized
    elif 'ebsoptimized' in usage_type_lower:
        return 'EC2 - EBS Optimized'

    # Dedicated Hosts
    elif 'dedicatedhost' in usage_type_lower or 'host' in usage_type_lower:
        return 'EC2 - Dedicated Hosts'

    # Spot Instances
    elif 'spot' in usage_type_lower:
        return 'EC2 - Spot Instances'

    # Reserved Instances
    elif 'reserved' in usage_type_lower or 'heavyusage' in usage_type_lower:
        return 'EC2 - Reserved Instances'

    # CloudWatch (métricas detalladas)
    elif 'cloudwatch' in usage_type_lower or 'monitoring' in usage_type_lower:
        return 'EC2 - CloudWatch Metrics'

    # Otros
    else:
        # Intentar mostrar un poco más de info del usage type
        if len(usage_type) > 40:
            return f'EC2 - Other ({usage_type[:40]}...)'
        else:
            return f'EC2 - Other ({usage_type})'


def crear_hoja_analisis(writer, datos_procesados):
    """
    Crea hoja de análisis por ServerGroup
    """
    from openpyxl.styles import Font, PatternFill

    # Análisis por ServerGroup
    analisis_servergroup = defaultdict(float)

    for clave_recurso, servicios in datos_procesados.items():
        partes = clave_recurso.split('|')
        servergroup = partes[1] if len(partes) > 1 and partes[1] else 'Sin ServerGroup'

        total = sum(servicios.values())
        analisis_servergroup[servergroup] += total

    # Crear hoja de análisis por ServerGroup
    if analisis_servergroup:
        datos_sg = [['ServerGroup', 'Costo Total (US$)']]
        for sg, costo in sorted(analisis_servergroup.items(), key=lambda x: x[1], reverse=True):
            datos_sg.append([sg, round(costo, 2)])

        df_sg = pd.DataFrame(datos_sg[1:], columns=datos_sg[0])
        df_sg.to_excel(writer, sheet_name='Por ServerGroup', index=False)

        ws_sg = writer.sheets['Por ServerGroup']
        ws_sg.column_dimensions['A'].width = 35
        ws_sg.column_dimensions['B'].width = 20

        # Formato del encabezado
        for cell in ws_sg[1]:
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)


def procesar_costos(costos_por_name, etiquetas_recursos, costos_backup, desglose_ec2):
    """
    Procesa los resultados y los organiza por etiquetas (Name, ServerGroup) y servicio
    Incluye el desglose detallado de EC2
    """
    datos_procesados = defaultdict(lambda: defaultdict(float))

    # Lista de servicios EC2 que serán reemplazados por el desglose
    servicios_ec2_a_reemplazar = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]

    # Procesar costos con sus etiquetas
    for name, servicios in costos_por_name.items():
        # Obtener etiquetas del recurso
        if name in etiquetas_recursos:
            servergroup = etiquetas_recursos[name].get('servergroup', '')
        else:
            servergroup = ''

        # Crear clave combinada con todas las etiquetas
        clave_recurso = f"{name}|{servergroup}"

        for servicio, costo in servicios.items():
            if costo > 0:
                # Si es un servicio EC2 que será desglosado y tenemos desglose para este recurso
                if servicio in servicios_ec2_a_reemplazar and name in desglose_ec2:
                    # NO agregar el servicio original, será reemplazado por el desglose
                    continue
                else:
                    # Agregar el servicio normalmente
                    datos_procesados[clave_recurso][servicio] += costo

        # Agregar los componentes desglosados de EC2 si existen
        if name in desglose_ec2:
            for servicio_detallado, costo_detallado in desglose_ec2[name].items():
                datos_procesados[clave_recurso][servicio_detallado] += costo_detallado

    # Agregar costos de AWS Backup
    for clave, costo in costos_backup.items():
        etiqueta_name, plan_backup = clave.split('|')
        servicio_backup = f"AWS Backup ({plan_backup})"

        # Buscar la clave del recurso que coincida con el Name
        clave_encontrada = None
        for clave_recurso in datos_procesados.keys():
            if clave_recurso.startswith(etiqueta_name + '|'):
                clave_encontrada = clave_recurso
                break

        if clave_encontrada:
            datos_procesados[clave_encontrada][servicio_backup] += costo
        else:
            # Si no se encuentra, crear nueva entrada
            clave_nueva = f"{etiqueta_name}|"
            datos_procesados[clave_nueva][servicio_backup] += costo

    return datos_procesados


def crear_excel(datos_procesados, fecha_inicio, fecha_fin, nombre_archivo):
    """
    Crea un archivo Excel con los datos de costos incluyendo Name y ServerGroup
    """
    # Crear lista de filas para el DataFrame
    filas = []

    for clave_recurso, servicios in sorted(datos_procesados.items()):
        # Separar las etiquetas
        partes = clave_recurso.split('|')
        etiqueta_name = partes[0]
        etiqueta_servergroup = partes[1] if len(partes) > 1 else ''

        total_recurso = sum(servicios.values())

        # Fila de resumen por recurso
        fila_resumen = {
            'Name': etiqueta_name,
            'ServerGroup': etiqueta_servergroup,
            'Servicio': 'TOTAL',
            'Costo (US$)': round(total_recurso, 2)
        }
        filas.append(fila_resumen)

        # Filas de detalle por servicio
        for servicio, costo in sorted(servicios.items(), key=lambda x: x[1], reverse=True):
            fila_detalle = {
                'Name': '',
                'ServerGroup': '',
                'Servicio': servicio,
                'Costo (US$)': round(costo, 2)
            }
            filas.append(fila_detalle)

        # Fila en blanco para separar
        filas.append({'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''})

    # Crear DataFrame
    df = pd.DataFrame(filas)

    # Calcular costo total general
    costo_total = sum(sum(servicios.values()) for servicios in datos_procesados.values())

    # Crear archivo Excel con formato
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Costos por Etiqueta', index=False)

        # Obtener el workbook y worksheet para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets['Costos por Etiqueta']

        # Ajustar anchos de columna
        worksheet.column_dimensions['A'].width = 35  # Name
        worksheet.column_dimensions['B'].width = 25  # ServerGroup
        worksheet.column_dimensions['C'].width = 50  # Servicio
        worksheet.column_dimensions['D'].width = 15  # Costo

        # Aplicar formato a las celdas de totales
        from openpyxl.styles import Font, PatternFill, Alignment

        fill_total = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font_bold = Font(bold=True)

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            if row[2].value == 'TOTAL':  # Columna "Servicio"
                for cell in row:
                    cell.fill = fill_total
                    cell.font = font_bold

        # Agregar hoja de resumen
        resumen = []
        resumen.append(['Periodo', f"{fecha_inicio} a {fecha_fin}"])
        resumen.append([''])
        resumen.append(['Name', 'ServerGroup', 'Costo Total (US$)'])

        for clave_recurso, servicios in sorted(datos_procesados.items(), key=lambda x: sum(x[1].values()),
                                               reverse=True):
            partes = clave_recurso.split('|')
            etiqueta_name = partes[0]
            etiqueta_servergroup = partes[1] if len(partes) > 1 else ''

            total = sum(servicios.values())
            resumen.append([etiqueta_name, etiqueta_servergroup, round(total, 2)])

        resumen.append([''])
        resumen.append(['COSTO TOTAL GENERAL', '', round(costo_total, 2)])

        df_resumen = pd.DataFrame(resumen)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)

        worksheet_resumen = writer.sheets['Resumen']
        worksheet_resumen.column_dimensions['A'].width = 35
        worksheet_resumen.column_dimensions['B'].width = 25
        worksheet_resumen.column_dimensions['C'].width = 20

        # Formato para el total general
        last_row = len(resumen) + 1
        for cell in worksheet_resumen[last_row]:
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            cell.font = Font(bold=True, size=12)

        # Agregar hoja de análisis por ServerGroup
        crear_hoja_analisis(writer, datos_procesados)

    print(f"\n✓ Archivo Excel creado: {nombre_archivo}")
    print(f"✓ Costo total del periodo: ${costo_total:,.2f} US$")
    return nombre_archivo


def main():
    parser = argparse.ArgumentParser(
        description='Extrae costos de AWS por etiqueta Name y exporta a Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
        Ejemplos de uso:
          %(prog)s                          # Costos del mes actual
          %(prog)s --mes 10 --anio 2024     # Costos de octubre 2024
          %(prog)s --mes 1 --anio 2025      # Costos de enero 2025
          %(prog)s --output costos_aws.xlsx # Especificar nombre de archivo
        """
    )

    parser.add_argument('--mes', type=int, help='Mes (1-12)')
    parser.add_argument('--anio', type=int, help='Año (ej: 2024)')
    parser.add_argument('--output', type=str, default='aws_costos_por_etiqueta.xlsx',
                        help='Nombre del archivo Excel de salida')
    parser.add_argument('--profile', type=str, help='Perfil de AWS CLI a usar')
    parser.add_argument('--region', type=str, default='us-east-1', help='Región de AWS')

    args = parser.parse_args()

    # Validar mes si se proporciona
    if args.mes and (args.mes < 1 or args.mes > 12):
        print("Error: El mes debe estar entre 1 y 12")
        sys.exit(1)

    # Validar que si se da mes, se dé año y viceversa
    if (args.mes and not args.anio) or (args.anio and not args.mes):
        print("Error: Debe especificar tanto --mes como --anio, o ninguno")
        sys.exit(1)

    print("=" * 60)
    print("AWS Cost Report - Costos por Etiqueta Name")
    print("=" * 60)

    # Obtener rango de fechas
    fecha_inicio, fecha_fin = obtener_rango_fechas(args.mes, args.anio)

    # Crear cliente de Cost Explorer
    try:
        session_params = {'region_name': args.region}
        if args.profile:
            session_params['profile_name'] = args.profile

        session = boto3.Session(**session_params)
        cliente_ce = session.client('ce')
        print(f"✓ Conectado a AWS (región: {args.region})")
    except Exception as e:
        print(f"Error al conectar con AWS: {e}")
        print("\nAsegúrese de tener configuradas sus credenciales de AWS")
        print("Puede usar: aws configure")
        sys.exit(1)

    # Obtener datos de costos
    costos_por_name, etiquetas_recursos = obtener_costos_por_servicio(cliente_ce, fecha_inicio, fecha_fin)
    costos_backup = obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin)
    desglose_ec2 = obtener_desglose_ec2(cliente_ce, fecha_inicio, fecha_fin)

    # Procesar datos
    print("Procesando datos...")
    datos_procesados = procesar_costos(costos_por_name, etiquetas_recursos, costos_backup, desglose_ec2)

    if not datos_procesados:
        print("\n⚠ No se encontraron costos en el periodo especificado")
        sys.exit(0)

    # Crear archivo Excel
    archivo_salida = crear_excel(datos_procesados, fecha_inicio, fecha_fin, args.output)

    print("\n" + "=" * 60)
    print(f"Recursos encontrados: {len(datos_procesados)}")
    print("=" * 60)


if __name__ == '__main__':
    main()