#!/usr/bin/env python3
"""
Script para extraer costos de AWS por etiqueta Name
Incluye todos los servicios (EC2, S3, RDS, etc.) y AWS Backup
Exporta los resultados a Excel
"""

import boto3
import pandas as pd
from datetime import datetime
from collections import defaultdict
import argparse
import sys


def obtener_rango_fechas(mes=None, anio=None):
    """
    Obtiene el rango de fechas para la consulta
    Si no se especifica mes/año, usa el mes actual
    """
    if mes and anio:
        fecha_inicio = datetime(anio, mes, 1)
    else:
        ahora = datetime.now()
        fecha_inicio = datetime(ahora.year, ahora.month, 1)

    if fecha_inicio.month == 12:
        fecha_fin = datetime(fecha_inicio.year + 1, 1, 1)
    else:
        fecha_fin = datetime(fecha_inicio.year, fecha_inicio.month + 1, 1)

    return fecha_inicio.strftime('%Y-%m-%d'), fecha_fin.strftime('%Y-%m-%d')


def obtener_costos_base(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene todos los costos agrupados por servicio y Name"""
    print("📊 Obteniendo costos base por Name y Servicio...")

    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'DIMENSION', 'Key': 'SERVICE'},
                {'Type': 'TAG', 'Key': 'Name'}
            ]
        )

        costos = defaultdict(lambda: defaultdict(float))
        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                servicio = grupo['Keys'][0]
                name = grupo['Keys'][1].replace('Name$', '') if grupo['Keys'][1] != 'Name$' else 'Sin etiqueta'
                costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                if costo > 0:
                    costos[name][servicio] += costo

        return costos
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


def obtener_servergroup(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene la etiqueta ServerGroup para cada Name"""
    print("🏷️  Obteniendo etiquetas ServerGroup...")

    servergroups = {}
    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'TAG', 'Key': 'Name'},
                {'Type': 'TAG', 'Key': 'ServerGroup'}
            ]
        )

        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][0] != 'Name$' else ''
                sg = grupo['Keys'][1].replace('ServerGroup$', '') if len(grupo['Keys']) > 1 and grupo['Keys'][
                    1] != 'ServerGroup$' else ''

                if name and name != 'Sin etiqueta' and sg:
                    servergroups[name] = sg

        return servergroups
    except Exception as e:
        print(f"⚠️  Advertencia ServerGroup: {e}")
        return {}


def obtener_desglose_ec2_completo(cliente_ce, fecha_inicio, fecha_fin, names_con_ec2):
    """Obtiene el desglose COMPLETO de EC2 por Usage Type - SOLO para Names que ya tienen EC2"""
    print("🔍 Desglosando EC2 en detalle...")

    desglose = defaultdict(lambda: defaultdict(float))

    # Todos los servicios relacionados con EC2
    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]

    for servicio in servicios_ec2:
        print(f"   → {servicio}")
        try:
            response = cliente_ce.get_cost_and_usage(
                TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
                Granularity='MONTHLY',
                Metrics=['UnblendedCost'],
                Filter={'Dimensions': {'Key': 'SERVICE', 'Values': [servicio]}},
                GroupBy=[
                    {'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'},
                    {'Type': 'TAG', 'Key': 'Name'}
                ]
            )

            for periodo in response['ResultsByTime']:
                for grupo in periodo['Groups']:
                    usage_type = grupo['Keys'][0]
                    name = grupo['Keys'][1].replace('Name$', '') if grupo['Keys'][1] != 'Name$' else 'Sin etiqueta'
                    costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                    # ✅ CRÍTICO: Solo agregar si este Name tiene EC2 en costos_base
                    # Esto evita capturar recursos sin etiqueta que AWS asocia automáticamente
                    if costo > 0 and name in names_con_ec2:
                        categoria = categorizar_usage_type(usage_type)
                        desglose[name][categoria] += costo

        except Exception as e:
            print(f"   ⚠️  {e}")
            continue

    return desglose


def categorizar_usage_type(usage_type):
    """Categoriza los Usage Types de EC2 en nombres descriptivos"""
    ut = usage_type.lower()

    # Instancias EC2
    if any(x in ut for x in ['boxusage', 'instanceusage', 'hoursusage']):
        # Extraer tipo de instancia si es posible
        if ':' in usage_type:
            tipo = usage_type.split(':')[-1]
            return f'EC2 - Instancia ({tipo})'
        return 'EC2 - Instancias'

    # EBS Volumes por tipo
    elif 'volumeusage' in ut:
        if 'gp2' in ut:
            return 'EC2 - EBS Volumes (gp2)'
        elif 'gp3' in ut:
            return 'EC2 - EBS Volumes (gp3)'
        elif 'io1' in ut or 'io2' in ut:
            return 'EC2 - EBS Volumes (io1/io2)'
        elif 'st1' in ut:
            return 'EC2 - EBS Volumes (st1)'
        elif 'sc1' in ut:
            return 'EC2 - EBS Volumes (sc1)'
        else:
            return 'EC2 - EBS Volumes'

    # Snapshots
    elif 'snapshot' in ut:
        return 'EC2 - EBS Snapshots'

    # IOPS provisionadas
    elif 'piops' in ut or 'volumeiops' in ut:
        return 'EC2 - EBS IOPS'

    # Throughput
    elif 'throughput' in ut:
        return 'EC2 - EBS Throughput'

    # Network Interfaces
    elif 'networkinterface' in ut or 'createnetworkinterface' in ut:
        return 'EC2 - Network Interfaces (ENI)'

    # Elastic IPs
    elif 'elasticip' in ut or 'idleaddress' in ut or 'addressusage' in ut:
        return 'EC2 - Elastic IPs'

    # Data Transfer
    elif 'datatransfer' in ut or 'data-transfer' in ut:
        if 'in-bytes' in ut or 'regional-bytes' in ut:
            return 'EC2 - Data Transfer (Regional/In)'
        elif 'out-bytes' in ut or 'bytes' in ut:
            return 'EC2 - Data Transfer (Out)'
        else:
            return 'EC2 - Data Transfer'

    # NAT Gateway
    elif 'natgateway' in ut:
        if 'bytes' in ut:
            return 'EC2 - NAT Gateway (Data Processed)'
        else:
            return 'EC2 - NAT Gateway (Hours)'

    # Load Balancers
    elif 'loadbalancer' in ut or 'elb:' in ut or 'lcu' in ut:
        if 'application' in ut or 'alb' in ut:
            return 'EC2 - Load Balancer (ALB)'
        elif 'network' in ut or 'nlb' in ut:
            return 'EC2 - Load Balancer (NLB)'
        else:
            return 'EC2 - Load Balancer'

    # VPN
    elif 'vpn' in ut:
        return 'EC2 - VPN Connection'

    # EBS Optimized
    elif 'ebsoptimized' in ut:
        return 'EC2 - EBS Optimized'

    # Spot Instances
    elif 'spot' in ut:
        return 'EC2 - Spot Instances'

    # CloudWatch
    elif 'cloudwatch' in ut or 'gmdetailedmonitoring' in ut:
        return 'EC2 - CloudWatch Monitoring'

    # Si no se puede categorizar, mostrar el usage type
    else:
        # Limpiar y acortar
        tipo_limpio = usage_type.replace('USE1-', '').replace('EUW1-', '')
        if len(tipo_limpio) > 50:
            return f'EC2 - {tipo_limpio[:50]}...'
        return f'EC2 - {tipo_limpio}'


def obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene costos de AWS Backup por Name (sin necesidad de etiqueta especial)"""
    print("💾 Obteniendo costos de AWS Backup...")

    backup_costs = defaultdict(float)

    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            Filter={'Dimensions': {'Key': 'SERVICE', 'Values': ['AWS Backup']}},
            GroupBy=[{'Type': 'TAG', 'Key': 'Name'}]
        )

        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][0] != 'Name$' else 'Sin etiqueta'
                costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                if costo > 0:
                    backup_costs[name] += costo

        return backup_costs
    except Exception as e:
        print(f"⚠️  Advertencia Backup: {e}")
        return {}


def normalizar_desglose_ec2(costos_base, desglose_ec2):
    """Normaliza el desglose EC2 para que coincida exactamente con costos_base por Name"""
    print("🔧 Normalizando desglose EC2...")

    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]

    desglose_normalizado = defaultdict(lambda: defaultdict(float))

    for name in desglose_ec2.keys():
        # Total EC2 en costos_base para este Name
        total_base = sum(costos_base.get(name, {}).get(s, 0) for s in servicios_ec2)

        # Total en desglose para este Name
        total_desglose = sum(desglose_ec2[name].values())

        if total_desglose > 0 and total_base > 0:
            # Factor de normalización
            factor = total_base / total_desglose

            # Aplicar factor a cada categoría
            for categoria, costo in desglose_ec2[name].items():
                desglose_normalizado[name][categoria] = costo * factor

            if abs(factor - 1.0) > 0.01:
                print(f"   ⚙️  {name}: factor={factor:.3f} (base=${total_base:.2f}, desglose=${total_desglose:.2f})")
        elif total_base > 0:
            # Hay costos en base pero no en desglose - mantener base sin desglosar
            print(f"   ⚠️  {name}: tiene EC2 en base (${total_base:.2f}) pero no en desglose")

    return desglose_normalizado


def diagnosticar_ec2(costos_base, desglose_ec2):
    """Diagnostica diferencias entre costos base de EC2 y desglose"""
    print("\n🔍 DIAGNÓSTICO DETALLADO DE EC2:")
    print("-" * 70)

    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]

    # Total EC2 en costos_base
    total_ec2_base = 0
    ec2_por_name = defaultdict(float)
    ec2_por_servicio = defaultdict(float)

    for name, servicios in costos_base.items():
        for servicio in servicios_ec2:
            if servicio in servicios:
                costo = servicios[servicio]
                total_ec2_base += costo
                ec2_por_name[name] += costo
                ec2_por_servicio[servicio] += costo

    # Total EC2 en desglose
    total_ec2_desglose = sum(sum(cats.values()) for cats in desglose_ec2.values())

    print(f"Total EC2 en costos_base: ${total_ec2_base:,.2f}")
    for servicio in servicios_ec2:
        print(f"  - {servicio}: ${ec2_por_servicio[servicio]:,.2f}")

    print(f"\nTotal EC2 en desglose: ${total_ec2_desglose:,.2f}")
    diferencia = total_ec2_base - total_ec2_desglose
    print(f"Diferencia: ${diferencia:,.2f}")

    if abs(diferencia) > 0.01:
        print(f"\n⚠️  ¡DIFERENCIA DE ${abs(diferencia):,.2f}!")

        # Names que tienen EC2 en base pero NO en desglose
        names_solo_base = set(ec2_por_name.keys()) - set(desglose_ec2.keys())
        if names_solo_base:
            total_sin_desglose = sum(ec2_por_name[n] for n in names_solo_base)
            print(f"\n⚠️  Names con EC2 en base pero SIN desglose ({len(names_solo_base)}):")
            print(f"    Total sin desglose: ${total_sin_desglose:,.2f}")
            for name in sorted(names_solo_base, key=lambda x: ec2_por_name[x], reverse=True)[:5]:
                print(f"  - {name}: ${ec2_por_name[name]:,.2f}")

        # Comparar totales por Name
        print(f"\n📊 Mayores diferencias por Name:")
        diferencias = []
        for name in set(ec2_por_name.keys()) | set(desglose_ec2.keys()):
            base = ec2_por_name[name]
            desg = sum(desglose_ec2.get(name, {}).values())
            if abs(base - desg) > 0.01:
                diferencias.append((name, base, desg, base - desg))

        for name, base, desg, diff in sorted(diferencias, key=lambda x: abs(x[3]), reverse=True)[:5]:
            print(f"  {name}: Base=${base:.2f}, Desglose=${desg:.2f}, Diff=${diff:.2f}")
    else:
        print(f"✅ Desglose EC2 completo y correcto")

    return total_ec2_base, total_ec2_desglose


def procesar_datos(costos_base, desglose_ec2, backup_costs, servergroups):
    """Procesa y combina todos los datos SIN DUPLICACIONES"""
    print("\n⚙️  Procesando datos...")

    datos_finales = defaultdict(lambda: {'servergroup': '', 'servicios': defaultdict(float)})

    # Servicios EC2 que serán reemplazados por el desglose
    servicios_ec2_a_reemplazar = {
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    }

    for name, servicios in costos_base.items():
        # Agregar ServerGroup
        datos_finales[name]['servergroup'] = servergroups.get(name, '')

        # Agregar servicios
        for servicio, costo in servicios.items():
            # ✅ CORRECCIÓN 1: Excluir AWS Backup si ya lo tenemos por separado
            if servicio == 'AWS Backup':
                # NO agregarlo aquí, lo agregaremos después desde backup_costs
                continue

            # ✅ CORRECCIÓN 2: Excluir servicios EC2 SOLO si tenemos desglose para este Name
            if servicio in servicios_ec2_a_reemplazar and name in desglose_ec2:
                # NO agregarlo, lo agregaremos desde el desglose
                continue

            # ✅ CORRECCIÓN 3: Si es un servicio EC2 pero NO tenemos desglose, SÍ agregarlo
            datos_finales[name]['servicios'][servicio] += costo

        # Agregar desglose de EC2 (solo si existe para este Name)
        if name in desglose_ec2:
            for categoria, costo in desglose_ec2[name].items():
                datos_finales[name]['servicios'][categoria] += costo

        # Agregar AWS Backup (solo si existe para este Name)
        if name in backup_costs:
            datos_finales[name]['servicios']['AWS Backup'] += backup_costs[name]

    # Verificación de totales
    total_procesado = sum(sum(info['servicios'].values()) for info in datos_finales.values())
    total_base = sum(sum(servicios.values()) for servicios in costos_base.values())

    print(f"Total en costos_base: ${total_base:,.2f}")
    print(f"Total procesado: ${total_procesado:,.2f}")

    if abs(total_procesado - total_base) > 1:
        print(f"⚠️  Diferencia en procesamiento: ${abs(total_procesado - total_base):,.2f}")
    else:
        print(f"✅ Procesamiento correcto (diferencia: ${abs(total_procesado - total_base):.2f})")

    return datos_finales


def crear_excel(datos, fecha_inicio, fecha_fin, nombre_archivo, es_partner=False, porcentaje_descuento=5.0):
    """Crea el archivo Excel con los resultados"""
    print("\n📝 Creando Excel...")

    # Calcular total general
    costo_total = sum(sum(info['servicios'].values()) for info in datos.values())

    # Calcular descuento si es partner
    monto_descuento = 0
    costo_con_descuento = costo_total
    if es_partner:
        monto_descuento = costo_total * (porcentaje_descuento / 100)
        costo_con_descuento = costo_total - monto_descuento

    filas = []

    # *** NUEVA SECCIÓN: TOTAL GENERAL AL INICIO ***
    filas.append({
        'Name': '*** TOTAL GENERAL ***',
        'ServerGroup': '',
        'Servicio': '',
        'Costo (US$)': round(costo_total, 2)
    })

    # Si es partner, añadir línea de descuento
    if es_partner:
        filas.append({
            'Name': f'Descuento Partner ({porcentaje_descuento}%)',
            'ServerGroup': '',
            'Servicio': '',
            'Costo (US$)': round(-monto_descuento, 2)
        })
        filas.append({
            'Name': '*** TOTAL CON DESCUENTO ***',
            'ServerGroup': '',
            'Servicio': '',
            'Costo (US$)': round(costo_con_descuento, 2)
        })

    # Línea en blanco separadora
    filas.append({'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''})
    filas.append({'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''})

    # Ordenar por costo total descendente
    datos_ordenados = sorted(
        datos.items(),
        key=lambda x: sum(x[1]['servicios'].values()),
        reverse=True
    )

    for name, info in datos_ordenados:
        total = sum(info['servicios'].values())
        servergroup = info['servergroup']

        # Fila de total
        filas.append({
            'Name': name,
            'ServerGroup': servergroup,
            'Servicio': '*** TOTAL ***',
            'Costo (US$)': round(total, 2)
        })

        # Servicios ordenados por costo
        for servicio, costo in sorted(info['servicios'].items(), key=lambda x: x[1], reverse=True):
            filas.append({
                'Name': '',
                'ServerGroup': '',
                'Servicio': servicio,
                'Costo (US$)': round(costo, 2)
            })

        # Línea en blanco
        filas.append({'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''})

    df = pd.DataFrame(filas)

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detalle de Costos', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Detalle de Costos']

        # Ajustar anchos
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 55
        worksheet.column_dimensions['D'].width = 15

        # Formato
        from openpyxl.styles import Font, PatternFill

        fill_total = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        fill_total_general = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        fill_descuento = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
        fill_total_descuento = PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid')  # Verde lima
        font_bold = Font(bold=True, size=11)
        font_bold_large = Font(bold=True, size=12)

        # Formatear filas
        for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            cell_value = row[0].value or ''
            servicio_value = row[2].value or ''

            # Total general al inicio
            if cell_value == '*** TOTAL GENERAL ***':
                for cell in row:
                    cell.fill = fill_total_general
                    cell.font = font_bold_large
            # Línea de descuento (en verde)
            elif 'Descuento Partner' in str(cell_value):
                for cell in row:
                    cell.fill = fill_descuento
                    cell.font = font_bold_large
            # Total con descuento (en verde más fuerte)
            elif cell_value == '*** TOTAL CON DESCUENTO ***':
                for cell in row:
                    cell.fill = fill_total_descuento
                    cell.font = font_bold_large
            # Totales de cada Name
            elif servicio_value == '*** TOTAL ***':
                for cell in row:
                    cell.fill = fill_total
                    cell.font = font_bold

        # Hoja de resumen
        resumen = [
            ['Periodo', f'{fecha_inicio} a {fecha_fin}'],
            [''],
            ['TOTAL GENERAL', '', round(costo_total, 2)]
        ]

        if es_partner:
            resumen.append([f'Descuento Partner ({porcentaje_descuento}%)', '', round(-monto_descuento, 2)])
            resumen.append(['TOTAL CON DESCUENTO', '', round(costo_con_descuento, 2)])

        resumen.append([''])
        resumen.append(['Name', 'ServerGroup', 'Costo Total (US$)'])

        for name, info in datos_ordenados:
            total = sum(info['servicios'].values())
            resumen.append([name, info['servergroup'], round(total, 2)])

        df_resumen = pd.DataFrame(resumen)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)

        ws_resumen = writer.sheets['Resumen']
        ws_resumen.column_dimensions['A'].width = 40
        ws_resumen.column_dimensions['B'].width = 25
        ws_resumen.column_dimensions['C'].width = 20

        # Formato en resumen
        # Total general
        for cell in ws_resumen[3]:
            cell.fill = fill_total_general
            cell.font = font_bold_large

        if es_partner:
            # Descuento
            for cell in ws_resumen[4]:
                cell.fill = fill_descuento
                cell.font = font_bold_large
            # Total con descuento
            for cell in ws_resumen[5]:
                cell.fill = fill_total_descuento
                cell.font = font_bold_large

        # Hoja por ServerGroup
        if any(info['servergroup'] for info in datos.values()):
            sg_totals = defaultdict(float)
            for name, info in datos.items():
                sg = info['servergroup'] if info['servergroup'] else 'Sin ServerGroup'
                sg_totals[sg] += sum(info['servicios'].values())

            sg_data = [['ServerGroup', 'Costo Total (US$)']]
            for sg, total in sorted(sg_totals.items(), key=lambda x: x[1], reverse=True):
                sg_data.append([sg, round(total, 2)])

            df_sg = pd.DataFrame(sg_data[1:], columns=sg_data[0])
            df_sg.to_excel(writer, sheet_name='Por ServerGroup', index=False)

            ws_sg = writer.sheets['Por ServerGroup']
            ws_sg.column_dimensions['A'].width = 35
            ws_sg.column_dimensions['B'].width = 20

            for cell in ws_sg[1]:
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)

    print(f"\n✅ Excel creado: {nombre_archivo}")
    print(f"💰 Costo total: ${costo_total:,.2f} USD")
    if es_partner:
        print(f"💚 Descuento ({porcentaje_descuento}%): ${monto_descuento:,.2f} USD")
        print(f"💰 Total con descuento: ${costo_con_descuento:,.2f} USD")
    print(f"📊 Recursos: {len(datos)}")

    return nombre_archivo


def main():
    parser = argparse.ArgumentParser(description='Extrae costos de AWS por Name con desglose EC2 completo')
    parser.add_argument('--mes', type=int, help='Mes (1-12)')
    parser.add_argument('--anio', type=int, help='Año')
    parser.add_argument('--output', type=str, default='aws_costos_detallados.xlsx', help='Archivo de salida')
    parser.add_argument('--profile', type=str, help='Perfil AWS')
    parser.add_argument('--region', type=str, default='eu-west-1', help='Región AWS')
    parser.add_argument('--partner', action='store_true', help='Aplicar descuento de partner')
    parser.add_argument('--descuento', type=float, default=5.0, help='Porcentaje de descuento (default: 5.0)')

    args = parser.parse_args()

    if (args.mes and not args.anio) or (args.anio and not args.mes):
        print("❌ Debes especificar mes Y año, o ninguno")
        sys.exit(1)

    print("=" * 70)
    print("AWS COST REPORT - Desglose Completo por Name")
    if args.partner:
        print(f"🤝 Modo Partner activado - Descuento: {args.descuento}%")
    print("=" * 70)

    # Obtener fechas
    fecha_inicio, fecha_fin = obtener_rango_fechas(args.mes, args.anio)

    # Cliente AWS
    session_params = {'region_name': args.region}
    if args.profile:
        session_params['profile_name'] = args.profile

    try:
        session = boto3.Session(**session_params)
        ce = session.client('ce')
        print(f"✅ Conectado a AWS ({args.region})")
    except Exception as e:
        print(f"❌ Error conectando: {e}")
        sys.exit(1)

    # Obtener datos
    costos_base = obtener_costos_base(ce, fecha_inicio, fecha_fin)
    servergroups = obtener_servergroup(ce, fecha_inicio, fecha_fin)

    # ✅ Calcular qué Names tienen EC2 en costos_base (para limitar el desglose)
    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]
    names_con_ec2 = set()
    for name, servicios in costos_base.items():
        if any(s in servicios for s in servicios_ec2):
            names_con_ec2.add(name)

    print(f"   → {len(names_con_ec2)} Names con costos EC2 detectados")

    desglose_ec2 = obtener_desglose_ec2_completo(ce, fecha_inicio, fecha_fin, names_con_ec2)
    backup_costs = obtener_costos_backup(ce, fecha_inicio, fecha_fin)

    # ✅ Normalizar el desglose para que coincida exactamente con costos_base
    desglose_ec2_normalizado = normalizar_desglose_ec2(costos_base, desglose_ec2)

    # DIAGNÓSTICO EC2 (después de normalizar)
    diagnosticar_ec2(costos_base, desglose_ec2_normalizado)

    # Procesar
    datos = procesar_datos(costos_base, desglose_ec2_normalizado, backup_costs, servergroups)

    if not datos:
        print("\n⚠️  No se encontraron costos")
        sys.exit(0)

    # Verificación final
    total_final = sum(sum(info['servicios'].values()) for info in datos.values())
    total_esperado = sum(sum(servicios.values()) for servicios in costos_base.values())

    print("\n" + "=" * 70)
    print("✅ VERIFICACIÓN FINAL:")
    print(f"   Total Cost Explorer esperado: ${total_esperado:,.2f}")
    print(f"   Total calculado: ${total_final:,.2f}")
    diferencia_final = abs(total_final - total_esperado)
    if diferencia_final < 1:
        print(f"   ✅ ¡COINCIDENCIA PERFECTA! (diff: ${diferencia_final:.2f})")
    else:
        print(f"   ⚠️  Diferencia: ${diferencia_final:,.2f}")
    print("=" * 70)

    # Crear Excel con información de partner
    crear_excel(datos, fecha_inicio, fecha_fin, args.output, args.partner, args.descuento)

    print("=" * 70)
    print("✨ Completado exitosamente")
    print("=" * 70)


if __name__ == '__main__':
    main()