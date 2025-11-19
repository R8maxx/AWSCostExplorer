#!/usr/bin/env python3
"""
Script para crear archivo Excel de ejemplo
Muestra la estructura de salida del script aws_cost_report_v2.py
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill

def crear_excel_ejemplo_v2():
    """Crea un Excel de ejemplo con la nueva estructura V2"""
    
    # Datos de ejemplo con desglose completo de EC2
    datos_detalle = [
        {'Name': 'produccion.avanza20rl.es', 'ServerGroup': 'PRL', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 834.28},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Instancia (t3.large)', 'Costo (US$)': 350.50},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Instancia (t3.medium)', 'Costo (US$)': 112.15},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Data Transfer (Out)', 'Costo (US$)': 142.97},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Volumes (gp3)', 'Costo (US$)': 85.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Network Interfaces (ENI)', 'Costo (US$)': 25.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Snapshots', 'Costo (US$)': 14.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Data Transfer (Regional/In)', 'Costo (US$)': 8.45},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'AWS Backup', 'Costo (US$)': 41.26},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon Simple Storage Service', 'Costo (US$)': 18.11},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon Virtual Private Cloud', 'Costo (US$)': 3.72},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon CloudWatch', 'Costo (US$)': 2.87},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'db-produccion.rds', 'ServerGroup': 'Database', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 520.45},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon Relational Database Service', 'Costo (US$)': 420.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Volumes (gp3)', 'Costo (US$)': 65.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Snapshots', 'Costo (US$)': 15.15},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'AWS Backup', 'Costo (US$)': 20.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'app-desarrollo', 'ServerGroup': 'WebServers', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 125.50},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Instancia (t2.micro)', 'Costo (US$)': 65.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Volumes (gp2)', 'Costo (US$)': 35.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Data Transfer (Out)', 'Costo (US$)': 15.20},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'AWS Backup', 'Costo (US$)': 10.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'vpc-principal', 'ServerGroup': 'Network', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 58.90},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - NAT Gateway (Hours)', 'Costo (US$)': 32.40},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - NAT Gateway (Data Processed)', 'Costo (US$)': 15.46},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon Virtual Private Cloud', 'Costo (US$)': 11.04},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'balanceador-web', 'ServerGroup': 'Network', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 42.80},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Load Balancer (ALB)', 'Costo (US$)': 35.50},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Data Transfer (Out)', 'Costo (US$)': 7.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'bucket-backups-s3', 'ServerGroup': 'Storage', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 95.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon Simple Storage Service', 'Costo (US$)': 85.20},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'AWS Backup', 'Costo (US$)': 10.10},
        {'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''},
        
        {'Name': 'Sin etiqueta', 'ServerGroup': '', 'Servicio': '*** TOTAL ***', 'Costo (US$)': 87.45},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - Instancia (t3.small)', 'Costo (US$)': 55.30},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'EC2 - EBS Volumes (gp2)', 'Costo (US$)': 18.00},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'AWS Lambda', 'Costo (US$)': 12.15},
        {'Name': '', 'ServerGroup': '', 'Servicio': 'Amazon CloudWatch', 'Costo (US$)': 2.00},
    ]
    
    df_detalle = pd.DataFrame(datos_detalle)
    
    # Datos de resumen
    datos_resumen = [
        ['Periodo', '2024-11-01 a 2024-12-01'],
        [''],
        ['Name', 'ServerGroup', 'Costo Total (US$)'],
        ['produccion.avanza20rl.es', 'PRL', 834.28],
        ['db-produccion.rds', 'Database', 520.45],
        ['app-desarrollo', 'WebServers', 125.50],
        ['bucket-backups-s3', 'Storage', 95.30],
        ['Sin etiqueta', '', 87.45],
        ['vpc-principal', 'Network', 58.90],
        ['balanceador-web', 'Network', 42.80],
        [''],
        ['*** TOTAL GENERAL ***', '', 1764.68]
    ]
    
    # Datos por ServerGroup
    datos_servergroup = [
        ['ServerGroup', 'Costo Total (US$)'],
        ['PRL', 834.28],
        ['Database', 520.45],
        ['WebServers', 125.50],
        ['Network', 101.70],
        ['Storage', 95.30],
        ['Sin ServerGroup', 87.45]
    ]
    
    nombre_archivo = 'aws_costos_ejemplo_v2.xlsx'
    
    print("📝 Creando archivo de ejemplo V2...")
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        # Hoja 1: Detalle de Costos
        df_detalle.to_excel(writer, sheet_name='Detalle de Costos', index=False)
        
        workbook = writer.book
        ws_detalle = writer.sheets['Detalle de Costos']
        
        # Ajustar anchos
        ws_detalle.column_dimensions['A'].width = 40
        ws_detalle.column_dimensions['B'].width = 25
        ws_detalle.column_dimensions['C'].width = 55
        ws_detalle.column_dimensions['D'].width = 15
        
        # Formato para totales
        fill_total = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        font_bold = Font(bold=True, size=11)
        
        for row in ws_detalle.iter_rows(min_row=2, max_row=ws_detalle.max_row):
            if row[2].value == '*** TOTAL ***':
                for cell in row:
                    cell.fill = fill_total
                    cell.font = font_bold
        
        # Hoja 2: Resumen
        df_resumen = pd.DataFrame(datos_resumen)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
        ws_resumen = writer.sheets['Resumen']
        ws_resumen.column_dimensions['A'].width = 40
        ws_resumen.column_dimensions['B'].width = 25
        ws_resumen.column_dimensions['C'].width = 20
        
        # Formato para total general
        last_row = len(datos_resumen) + 1
        for cell in ws_resumen[last_row]:
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            cell.font = Font(bold=True, size=12)
        
        # Hoja 3: Por ServerGroup
        df_sg = pd.DataFrame(datos_servergroup[1:], columns=datos_servergroup[0])
        df_sg.to_excel(writer, sheet_name='Por ServerGroup', index=False)
        
        ws_sg = writer.sheets['Por ServerGroup']
        ws_sg.column_dimensions['A'].width = 35
        ws_sg.column_dimensions['B'].width = 20
        
        # Formato encabezado
        for cell in ws_sg[1]:
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
    
    print(f"✅ Archivo creado: {nombre_archivo}")
    print(f"💰 Total de ejemplo: $1,764.68 USD")
    print(f"📊 Recursos: 7")
    print(f"")
    print(f"📋 Este archivo muestra:")
    print(f"   • Desglose COMPLETO de EC2 (sin 'Others')")
    print(f"   • Tipos específicos de instancias y volúmenes")
    print(f"   • Network Interfaces identificadas")
    print(f"   • AWS Backup automático por Name")
    print(f"   • ServerGroup como información adicional")
    print(f"")
    print(f"🎯 Ejemplo de lo que verás en tu reporte real")

if __name__ == '__main__':
    crear_excel_ejemplo_v2()