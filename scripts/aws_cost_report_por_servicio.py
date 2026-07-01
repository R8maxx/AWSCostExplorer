#!/usr/bin/env python3
"""
AWS Cost Report - Una hoja por servicio (con estilos, filtros y gráficas)
=========================================================================
Genera un Excel con:
  - Hoja "Resumen": total general, tabla por servicio, Top Names y GRÁFICAS
    (barras por servicio, tarta de composición, barras Top-15 por Name)
  - Hoja "EC2": Compute + EC2-Other + EBS fusionados y desglosados por Usage
    Type y Name (Name repetido en cada fila para poder filtrar)
  - Una hoja por cada servicio principal (S3, RDS, Backup, CloudWatch, ...)
    + cualquier servicio que supere el umbral de coste
  - Hoja "Otros servicios": el resto de servicios agrupados

Cada hoja incluye:
  - Título y descripción del servicio (para quien no sepa qué es)
  - AutoFiltro para buscar por Name
  - Estilos: cabeceras, filas alternas, formato moneda, cabeceras fijas

Reutiliza la lógica de extracción y desglose EC2 ya validada en aws_cost_report.py
para que el total reconcilie exactamente con Cost Explorer.
"""

import boto3
from collections import defaultdict
import argparse
import hashlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties

from aws_cost_report import (
    obtener_rango_fechas,
    obtener_costos_base,
    obtener_desglose_ec2_completo,
    normalizar_desglose_ec2,
)

# --------------------------------------------------------------------------
# Configuración de servicios
# --------------------------------------------------------------------------
SERVICIOS_EC2 = [
    'Amazon Elastic Compute Cloud - Compute',
    'EC2 - Other',
    'Amazon Elastic Block Store',
]

PRINCIPALES = {
    'Amazon Simple Storage Service',
    'Amazon Relational Database Service',
    'AWS Backup',
    'Amazon CloudWatch',
    'AmazonCloudWatch',
    'Amazon Route 53',
    'Amazon Elastic Load Balancing',
    'Amazon Virtual Private Cloud',
    'Amazon Bedrock',
}

NOMBRES_HOJA = {
    'Amazon Simple Storage Service': 'S3',
    'Amazon Relational Database Service': 'RDS',
    'AWS Backup': 'Backup',
    'Amazon CloudWatch': 'CloudWatch',
    'AmazonCloudWatch': 'CloudWatch',
    'Amazon Route 53': 'Route 53',
    'Amazon Elastic Load Balancing': 'ELB',
    'Amazon Virtual Private Cloud': 'VPC',
    'Amazon Bedrock': 'Bedrock',
    'Amazon OpenSearch Service': 'OpenSearch',
    'Amazon Glacier': 'Glacier',
    'AWS WAF': 'WAF',
    'Amazon Simple Email Service': 'SES',
    'Amazon Simple Notification Service': 'SNS',
    'Amazon Simple Queue Service': 'SQS',
    'AWS Lambda': 'Lambda',
    'Amazon Elastic Container Service': 'ECS',
    'Amazon Elastic Container Registry (ECR)': 'ECR',
    'Amazon Elastic File System': 'EFS',
    'AWS Key Management Service': 'KMS',
    'Amazon DynamoDB': 'DynamoDB',
    'AWS Cost Explorer': 'Cost Explorer',
    'Tax': 'Impuestos',
}

# Descripciones legibles (para quien no conozca el servicio)
DESCRIPCIONES = {
    'EC2': ('Amazon EC2 — servidores virtuales (instancias) en la nube. Aquí se incluye el tiempo de '
            'cómputo, los discos EBS, snapshots, transferencia de datos, IPs elásticas y NAT.'),
    'Amazon Simple Storage Service': ('Amazon S3 — almacenamiento de objetos (archivos, backups, estáticos '
                                       'web). Se factura por GB almacenado, peticiones y transferencia.'),
    'Amazon Relational Database Service': ('Amazon RDS — bases de datos relacionales gestionadas (MySQL, '
                                           'PostgreSQL, etc.): cómputo, almacenamiento y backups de la BD.'),
    'AWS Backup': ('AWS Backup — copias de seguridad gestionadas por el servicio AWS Backup (EBS, RDS, EFS...); '
                   'se factura por el almacenamiento en el vault. OJO: los snapshots de EBS manuales o por '
                   'lifecycle (DLM) NO aparecen aquí, sino en la hoja EC2 de cada recurso como '
                   '"EC2 - EBS Snapshots". Esta hoja solo refleja las copias del servicio AWS Backup.'),
    'Amazon CloudWatch': 'Amazon CloudWatch — monitorización, métricas, logs y alarmas de los recursos AWS.',
    'AmazonCloudWatch': 'Amazon CloudWatch — monitorización, métricas, logs y alarmas de los recursos AWS.',
    'Amazon Virtual Private Cloud': ('Amazon VPC — red privada virtual. Incluye NAT Gateways, endpoints, '
                                     'IPs públicas y la transferencia de datos asociada.'),
    'Amazon Elastic Load Balancing': ('Elastic Load Balancing — balanceadores de carga (ALB/NLB) que reparten '
                                       'el tráfico entrante entre varias instancias.'),
    'Amazon OpenSearch Service': ('Amazon OpenSearch — motor de búsqueda y analítica de logs '
                                  '(antiguo Elasticsearch Service).'),
    'Amazon Glacier': 'Amazon S3 Glacier — almacenamiento de archivado a largo plazo de muy bajo coste.',
    'AWS WAF': ('AWS WAF — firewall de aplicaciones web que protege frente a ataques (SQLi, XSS, bots...).'),
    'Tax': 'Impuestos — IVA u otros impuestos aplicados por AWS sobre la factura.',
    'Amazon Route 53': 'Amazon Route 53 — DNS gestionado y registro de dominios.',
    'Amazon Bedrock': 'Amazon Bedrock — modelos de IA generativa gestionados (LLMs) accesibles vía API.',
    'AWS Lambda': 'AWS Lambda — ejecución de código sin servidores (serverless); se paga por uso.',
    'AWS Key Management Service': 'AWS KMS — gestión de claves de cifrado.',
    'Otros': ('Servicios de AWS con un coste individual por debajo del umbral, agrupados aquí. '
              'Filtra por la columna Servicio o Name.'),
}

# --------------------------------------------------------------------------
# Paleta y estilos
# --------------------------------------------------------------------------
C_TINTA   = '232F3E'  # azul marino AWS (squid ink)
C_NARANJA = 'FF9900'  # naranja AWS
C_AZUL    = '146EB4'  # azul AWS
C_BANDA   = 'EAF1F8'  # azul muy claro para filas alternas
C_BLANCO  = 'FFFFFF'
C_GOLD    = 'FFF2CC'  # dorado suave para subtotales
C_VERDE   = '2E7D32'  # verde descuento
C_VERDE_CL= 'D5F5E3'
C_GRIS    = 'F2F4F7'

PALETA_GRAFICA = ['FF9900', '146EB4', '232F3E', '2E7D32', 'C7511F', '7D3C98',
                  '16A085', 'D4AC0D', 'CB4335', '5D6D7E', '2874A6', 'A04000',
                  '117A65', '6C3483', '922B21']

CUR = '"$"#,##0.00'
_THIN = Side(style='thin', color='D9DEE3')
BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

F_TITULO   = Font(bold=True, size=18, color=C_BLANCO, name='Calibri')
F_DESC     = Font(italic=True, size=10, color='44546A', name='Calibri')
F_HEADER   = Font(bold=True, size=11, color=C_BLANCO, name='Calibri')
F_KPI_LBL  = Font(bold=True, size=12, color=C_BLANCO, name='Calibri')
F_KPI_VAL  = Font(bold=True, size=14, color=C_TINTA, name='Calibri')
F_SUBTOTAL = Font(bold=True, size=11, color=C_TINTA, name='Calibri')
F_NORMAL   = Font(size=10, name='Calibri')

FILL_TITULO = PatternFill('solid', fgColor=C_TINTA)
FILL_HEADER = PatternFill('solid', fgColor=C_AZUL)
FILL_BANDA  = PatternFill('solid', fgColor=C_BANDA)
FILL_BLANCO = PatternFill('solid', fgColor=C_BLANCO)
FILL_GOLD   = PatternFill('solid', fgColor=C_GOLD)
FILL_KPI    = PatternFill('solid', fgColor=C_NARANJA)
FILL_DESC   = PatternFill('solid', fgColor=C_GRIS)
FILL_VERDE  = PatternFill('solid', fgColor=C_VERDE_CL)


# Color FIJO por servicio (clave = nombre del servicio en Cost Explorer, o 'EC2').
# Se asigna por servicio, NO por posición, para que sea idéntico mes a mes.
COLOR_SERVICIO = {
    'EC2': '146EB4',                                   # azul AWS
    'Amazon Relational Database Service': 'C7511F',    # naranja quemado
    'Amazon Simple Storage Service': '2E7D32',         # verde
    'AWS Backup': '117A65',                            # teal
    'AmazonCloudWatch': '1F618D',                      # azul oscuro
    'Amazon CloudWatch': '1F618D',
    'Amazon Virtual Private Cloud': 'CB4335',          # rojo
    'Amazon Elastic Load Balancing': '148F77',         # verde azulado
    'Amazon OpenSearch Service': 'B7950B',             # mostaza
    'Amazon Glacier': 'AF601A',                        # ámbar
    'AWS WAF': '6C3483',                               # violeta
    'Tax': '7D3C98',                                   # púrpura
    'Amazon Route 53': '922B21',                       # granate
    'Amazon Bedrock': '5B2C6F',                        # violeta oscuro
    'AWS Lambda': '784212',                            # marrón
    'Amazon DynamoDB': '1A5276',                       # azul marino
    'AWS Key Management Service': 'A93226',            # rojo oscuro
    'Amazon Simple Email Service': '196F3D',           # verde oscuro
    'Amazon Simple Notification Service': '5D4037',    # marrón grisáceo
    'Amazon Simple Queue Service': '00695C',           # teal oscuro
    'Amazon Elastic File System': '4A148C',            # púrpura intenso
    'Amazon Elastic Container Service': '0D47A1',      # azul intenso
    'Amazon Route 53 Resolver': '827717',              # oliva
}
# Paleta de reserva para servicios sin color fijo (elegida de forma determinista por hash)
COLORES_FALLBACK = ['2874A6', '9A7D0A', '7B241C', '1E8449', '5B2C6F', '935116',
                    '148F77', '512E5F', '6E2C00', '154360', '7D6608', '4A235A']
COLOR_OTROS = '5D6D7E'  # gris azulado neutro para "Otros servicios"


def color_de_servicio(servicio):
    """Devuelve un color HEX FIJO para el servicio. Estable entre ejecuciones/meses:
    si no está en el mapa, se deriva de forma determinista del nombre (hash md5)."""
    if servicio in COLOR_SERVICIO:
        return COLOR_SERVICIO[servicio]
    idx = int(hashlib.md5(servicio.encode('utf-8')).hexdigest(), 16) % len(COLORES_FALLBACK)
    return COLORES_FALLBACK[idx]


def descripcion(servicio):
    if servicio in DESCRIPCIONES:
        return DESCRIPCIONES[servicio]
    return f'Servicio de AWS: {servicio}.'


def _etiquetas(showVal=False, showPercent=False):
    """DataLabelList mostrando SOLO lo indicado (evita el amontonamiento de LibreOffice)."""
    dl = DataLabelList()
    dl.showVal = showVal
    dl.showPercent = showPercent
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    return dl


# --------------------------------------------------------------------------
# Reorganización de datos
# --------------------------------------------------------------------------
def reorganizar_por_servicio(costos_base):
    servicios_data = defaultdict(lambda: defaultdict(float))
    for name, servicios in costos_base.items():
        for servicio, costo in servicios.items():
            if servicio in SERVICIOS_EC2:
                continue
            servicios_data[servicio][name] += costo
    return servicios_data


def clasificar_servicios(servicios_data, umbral):
    con_hoja = {}
    otros = defaultdict(lambda: defaultdict(float))
    for servicio, names in servicios_data.items():
        total = sum(names.values())
        if servicio in PRINCIPALES or total >= umbral:
            con_hoja[servicio] = dict(names)
        else:
            for n, c in names.items():
                otros[servicio][n] += c
    return con_hoja, otros


def nombre_hoja(servicio, usados):
    base = NOMBRES_HOJA.get(servicio, servicio)
    for ch in '[]:*?/\\':
        base = base.replace(ch, ' ')
    base = base.strip()[:31] or 'Servicio'
    candidato, i = base, 2
    while candidato in usados:
        sufijo = f' ({i})'
        candidato = base[:31 - len(sufijo)] + sufijo
        i += 1
    usados.add(candidato)
    return candidato


# --------------------------------------------------------------------------
# Helpers de estilo
# --------------------------------------------------------------------------
def _merge_estilo(ws, rango, valor, fill=None, font=None, align=None):
    ws.merge_cells(rango)
    top_left = rango.split(':')[0]
    ws[top_left] = valor
    if font:
        ws[top_left].font = font
    if align:
        ws[top_left].alignment = align
    if fill:
        # aplicar fill a todas las celdas del rango combinado
        for fila in ws[rango]:
            for c in fila:
                c.fill = fill


def _cabecera_hoja(ws, titulo, desc, total, ncols, color=None):
    """Escribe título + descripción + KPI de total. Devuelve la fila donde empieza la tabla.
    Si se pasa `color`, colorea el título y la pestaña de la hoja con ese color."""
    fill_titulo = PatternFill('solid', fgColor=color) if color else FILL_TITULO
    if color:
        ws.sheet_properties.tabColor = color
    ultima = get_column_letter(ncols)
    # Fila 1: título
    _merge_estilo(ws, f'A1:{ultima}1', titulo, fill_titulo, F_TITULO,
                  Alignment(horizontal='left', vertical='center', indent=1))
    ws.row_dimensions[1].height = 30
    # Filas 2-3: descripción
    _merge_estilo(ws, f'A2:{ultima}3', desc, FILL_DESC, F_DESC,
                  Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1))
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    # Fila 4: KPI total
    _merge_estilo(ws, f'A4:{get_column_letter(ncols-1)}4', 'TOTAL DEL SERVICIO',
                  FILL_KPI, F_KPI_LBL, Alignment(horizontal='right', vertical='center'))
    cell = ws.cell(4, ncols, round(total, 2))
    cell.fill = FILL_KPI
    cell.font = F_KPI_VAL
    cell.number_format = CUR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 22
    return 6  # la tabla (cabecera) empieza en la fila 6


def _formato_columnas(ws, anchos):
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho


# --------------------------------------------------------------------------
# Hojas
# --------------------------------------------------------------------------
def escribir_hoja_servicio(wb, hoja, servicio, datos, total, color):
    """Hoja simple: Name | Costo, con filtro y estilo."""
    ws = wb.create_sheet(hoja)
    _formato_columnas(ws, {'A': 48, 'B': 18})
    h = _cabecera_hoja(ws, NOMBRES_HOJA.get(servicio, servicio), descripcion(servicio), total, 2, color)
    fill_header = PatternFill('solid', fgColor=color)

    encabezados = ['Name', 'Costo (US$)']
    for c, texto in enumerate(encabezados, start=1):
        cell = ws.cell(h, c, texto)
        cell.fill = fill_header
        cell.font = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDE

    filas = sorted(datos.items(), key=lambda x: x[1], reverse=True)
    for i, (name, costo) in enumerate(filas):
        r = h + 1 + i
        fill = FILL_BANDA if i % 2 else FILL_BLANCO
        cn = ws.cell(r, 1, name); cn.fill = fill; cn.font = F_NORMAL; cn.border = BORDE
        cc = ws.cell(r, 2, round(costo, 2)); cc.fill = fill; cc.font = F_NORMAL
        cc.number_format = CUR; cc.border = BORDE
        cc.alignment = Alignment(horizontal='right')

    ultima = h + len(filas)
    ws.auto_filter.ref = f'A{h}:B{ultima}'
    ws.freeze_panes = f'A{h + 1}'


def escribir_hoja_ec2(wb, ec2_data, total, color):
    ws = wb.create_sheet('EC2')
    _formato_columnas(ws, {'A': 40, 'B': 46, 'C': 16})
    h = _cabecera_hoja(ws, 'EC2', DESCRIPCIONES['EC2'], total, 3, color)
    fill_header = PatternFill('solid', fgColor=color)

    for c, texto in enumerate(['Name', 'Detalle', 'Costo (US$)'], start=1):
        cell = ws.cell(h, c, texto)
        cell.fill = fill_header
        cell.font = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDE

    r = h + 1
    grupos = sorted(ec2_data.items(), key=lambda x: sum(x[1].values()), reverse=True)
    for gi, (name, cats) in enumerate(grupos):
        # Fila subtotal del Name (dorada, en negrita)
        for c, val in enumerate([name, '▸ TOTAL', round(sum(cats.values()), 2)], start=1):
            cell = ws.cell(r, c, val)
            cell.fill = FILL_GOLD
            cell.font = F_SUBTOTAL
            cell.border = BORDE
            if c == 3:
                cell.number_format = CUR
                cell.alignment = Alignment(horizontal='right')
        r += 1
        # Categorías (banda por grupo)
        fill = FILL_BANDA if gi % 2 else FILL_BLANCO
        for cat, costo in sorted(cats.items(), key=lambda x: x[1], reverse=True):
            cn = ws.cell(r, 1, name); cn.fill = fill; cn.font = F_NORMAL; cn.border = BORDE
            cd = ws.cell(r, 2, cat); cd.fill = fill; cd.font = F_NORMAL; cd.border = BORDE
            cc = ws.cell(r, 3, round(costo, 2)); cc.fill = fill; cc.font = F_NORMAL
            cc.number_format = CUR; cc.border = BORDE
            cc.alignment = Alignment(horizontal='right')
            r += 1

    ws.auto_filter.ref = f'A{h}:C{r - 1}'
    ws.freeze_panes = f'A{h + 1}'


def escribir_hoja_otros(wb, otros, total, color):
    ws = wb.create_sheet('Otros servicios')
    _formato_columnas(ws, {'A': 42, 'B': 42, 'C': 16})
    h = _cabecera_hoja(ws, 'Otros servicios', DESCRIPCIONES['Otros'], total, 3, color)
    fill_header = PatternFill('solid', fgColor=color)

    for c, texto in enumerate(['Servicio', 'Name', 'Costo (US$)'], start=1):
        cell = ws.cell(h, c, texto)
        cell.fill = fill_header
        cell.font = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDE

    r = h + 1
    grupos = sorted(otros.items(), key=lambda kv: sum(kv[1].values()), reverse=True)
    for gi, (servicio, names) in enumerate(grupos):
        for c, val in enumerate([servicio, '▸ TOTAL', round(sum(names.values()), 2)], start=1):
            cell = ws.cell(r, c, val)
            cell.fill = FILL_GOLD
            cell.font = F_SUBTOTAL
            cell.border = BORDE
            if c == 3:
                cell.number_format = CUR
                cell.alignment = Alignment(horizontal='right')
        r += 1
        fill = FILL_BANDA if gi % 2 else FILL_BLANCO
        for name, costo in sorted(names.items(), key=lambda x: x[1], reverse=True):
            cs = ws.cell(r, 1, servicio); cs.fill = fill; cs.font = F_NORMAL; cs.border = BORDE
            cn = ws.cell(r, 2, name); cn.fill = fill; cn.font = F_NORMAL; cn.border = BORDE
            cc = ws.cell(r, 3, round(costo, 2)); cc.fill = fill; cc.font = F_NORMAL
            cc.number_format = CUR; cc.border = BORDE
            cc.alignment = Alignment(horizontal='right')
            r += 1

    ws.auto_filter.ref = f'A{h}:C{r - 1}'
    ws.freeze_panes = f'A{h + 1}'


def escribir_hoja_resumen(wb, totales_servicio, totales_name, fecha_inicio, fecha_fin,
                          costo_total, es_partner, porcentaje_descuento):
    ws = wb.active
    ws.title = 'Resumen'
    _formato_columnas(ws, {'A': 34, 'B': 18, 'C': 3})

    # Título
    _merge_estilo(ws, 'A1:B1', 'AWS · Informe de costes', FILL_TITULO, F_TITULO,
                  Alignment(horizontal='left', vertical='center', indent=1))
    ws.row_dimensions[1].height = 32
    _merge_estilo(ws, 'A2:B2', f'Periodo: {fecha_inicio} a {fecha_fin}', FILL_DESC, F_DESC,
                  Alignment(horizontal='left', vertical='center', indent=1))

    # KPIs
    _merge_estilo(ws, 'A4:A4', 'TOTAL GENERAL', FILL_KPI, F_KPI_LBL,
                  Alignment(horizontal='right', vertical='center'))
    kpi = ws.cell(4, 2, round(costo_total, 2))
    kpi.fill = FILL_KPI; kpi.font = F_KPI_VAL; kpi.number_format = CUR
    kpi.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 24

    fila = 5
    if es_partner:
        monto = costo_total * (porcentaje_descuento / 100)
        _merge_estilo(ws, f'A{fila}:A{fila}', f'Descuento Partner ({porcentaje_descuento}%)',
                      FILL_VERDE, F_SUBTOTAL, Alignment(horizontal='right', vertical='center'))
        d = ws.cell(fila, 2, round(-monto, 2)); d.fill = FILL_VERDE; d.font = F_SUBTOTAL
        d.number_format = CUR; d.alignment = Alignment(horizontal='center')
        fila += 1
        _merge_estilo(ws, f'A{fila}:A{fila}', 'TOTAL CON DESCUENTO', PatternFill('solid', fgColor=C_VERDE),
                      Font(bold=True, size=12, color=C_BLANCO),
                      Alignment(horizontal='right', vertical='center'))
        td = ws.cell(fila, 2, round(costo_total - monto, 2))
        td.fill = PatternFill('solid', fgColor=C_VERDE); td.font = Font(bold=True, size=12, color=C_BLANCO)
        td.number_format = CUR; td.alignment = Alignment(horizontal='center')
        fila += 1

    # ---- Tabla: coste por servicio ----
    hs = fila + 1  # fila de cabecera de la tabla de servicios
    for c, texto in enumerate(['Servicio', 'Coste (US$)'], start=1):
        cell = ws.cell(hs, c, texto); cell.fill = FILL_HEADER; cell.font = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = BORDE
    for i, (etiqueta, total) in enumerate(totales_servicio):
        r = hs + 1 + i
        f = FILL_BANDA if i % 2 else FILL_BLANCO
        a = ws.cell(r, 1, etiqueta); a.fill = f; a.font = F_NORMAL; a.border = BORDE
        b = ws.cell(r, 2, round(total, 2)); b.fill = f; b.font = F_NORMAL
        b.number_format = CUR; b.border = BORDE; b.alignment = Alignment(horizontal='right')
    fin_serv = hs + len(totales_servicio)
    ws.auto_filter.ref = f'A{hs}:B{fin_serv}'

    # ---- Tabla: Top Names ----
    top_names = totales_name[:15]
    hn = fin_serv + 3
    _merge_estilo(ws, f'A{hn - 1}:B{hn - 1}', 'Top 15 recursos por coste (Name)',
                  FILL_DESC, F_SUBTOTAL, Alignment(horizontal='left', indent=1))
    for c, texto in enumerate(['Name', 'Coste (US$)'], start=1):
        cell = ws.cell(hn, c, texto); cell.fill = FILL_HEADER; cell.font = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = BORDE
    for i, (name, total) in enumerate(top_names):
        r = hn + 1 + i
        f = FILL_BANDA if i % 2 else FILL_BLANCO
        a = ws.cell(r, 1, name); a.fill = f; a.font = F_NORMAL; a.border = BORDE
        b = ws.cell(r, 2, round(total, 2)); b.fill = f; b.font = F_NORMAL
        b.number_format = CUR; b.border = BORDE; b.alignment = Alignment(horizontal='right')
    fin_name = hn + len(top_names)

    # ---- Gráficas ----
    # 1) Barras: coste por servicio
    bar = BarChart()
    bar.type = 'bar'
    bar.title = 'Coste por servicio (US$)'
    bar.height = 9
    bar.width = 20
    bar.legend = None
    data = Reference(ws, min_col=2, min_row=hs, max_row=fin_serv)
    cats = Reference(ws, min_col=1, min_row=hs + 1, max_row=fin_serv)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.dataLabels = _etiquetas(showVal=True)
    bar.series[0].graphicalProperties = GraphicalProperties(solidFill=C_AZUL)
    ws.add_chart(bar, 'D4')

    # 2) Tarta: composición por servicio
    pie = PieChart()
    pie.title = 'Composición del gasto por servicio'
    pie.height = 9
    pie.width = 12
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = _etiquetas(showPercent=True)
    for i in range(len(totales_servicio)):
        color = PALETA_GRAFICA[i % len(PALETA_GRAFICA)]
        pt = DataPoint(idx=i)
        pt.graphicalProperties = GraphicalProperties(solidFill=color)
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, 'D23')

    # 3) Barras: Top Names
    bar2 = BarChart()
    bar2.type = 'bar'
    bar2.title = 'Top 15 recursos por coste (Name)'
    bar2.height = 10
    bar2.width = 20
    bar2.legend = None
    data2 = Reference(ws, min_col=2, min_row=hn, max_row=fin_name)
    cats2 = Reference(ws, min_col=1, min_row=hn + 1, max_row=fin_name)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.dataLabels = _etiquetas(showVal=True)
    bar2.series[0].graphicalProperties = GraphicalProperties(solidFill=C_NARANJA)
    ws.add_chart(bar2, 'K4')


def crear_excel(ec2_data, con_hoja, otros, totales_name, fecha_inicio, fecha_fin,
                nombre_archivo, es_partner=False, porcentaje_descuento=5.0):
    print("\n📝 Creando Excel por servicio (con estilos y gráficas)...")

    ec2_total = sum(sum(cats.values()) for cats in ec2_data.values())
    totales_servicio = []
    if ec2_data:
        totales_servicio.append(('EC2 (Compute + Other + EBS)', ec2_total))
    for servicio, names in con_hoja.items():
        totales_servicio.append((NOMBRES_HOJA.get(servicio, servicio), sum(names.values())))
    otros_total = sum(sum(n.values()) for n in otros.values())
    if otros_total > 0:
        totales_servicio.append(('Otros servicios', otros_total))
    totales_servicio.sort(key=lambda x: x[1], reverse=True)
    costo_total = sum(t for _, t in totales_servicio)

    wb = Workbook()

    # Resumen (usa la hoja activa) — pestaña en azul marino corporativo
    escribir_hoja_resumen(wb, totales_servicio, totales_name, fecha_inicio, fecha_fin,
                          costo_total, es_partner, porcentaje_descuento)
    wb.active.sheet_properties.tabColor = C_TINTA

    # EC2 (color fijo)
    escribir_hoja_ec2(wb, ec2_data, ec2_total, color_de_servicio('EC2'))

    # Servicios con hoja propia (orden por total desc), cada uno con su color FIJO
    usados = {'Resumen', 'EC2'}
    servicios_ordenados = sorted(con_hoja, key=lambda s: sum(con_hoja[s].values()), reverse=True)
    for servicio in servicios_ordenados:
        hoja = nombre_hoja(servicio, usados)
        escribir_hoja_servicio(wb, hoja, servicio, con_hoja[servicio],
                               sum(con_hoja[servicio].values()), color_de_servicio(servicio))

    # Otros (color neutro)
    if otros_total > 0:
        escribir_hoja_otros(wb, otros, otros_total, COLOR_OTROS)

    wb.save(nombre_archivo)

    print(f"\n✅ Excel creado: {nombre_archivo}")
    print(f"💰 Costo total: ${costo_total:,.2f} USD")
    if es_partner:
        monto = costo_total * (porcentaje_descuento / 100)
        print(f"💚 Descuento ({porcentaje_descuento}%): ${monto:,.2f} USD")
        print(f"💰 Total con descuento: ${costo_total - monto:,.2f} USD")
    print(f"📄 Hojas: Resumen + EC2 + {len(con_hoja)} servicios" + (" + Otros" if otros_total > 0 else ""))
    return costo_total


def main():
    parser = argparse.ArgumentParser(description='Costos AWS con una hoja por servicio (EC2 desglosado)')
    parser.add_argument('--mes', type=int, help='Mes (1-12)')
    parser.add_argument('--anio', type=int, help='Año')
    parser.add_argument('--output', type=str, default='aws_costos_por_servicio.xlsx', help='Archivo de salida')
    parser.add_argument('--profile', type=str, help='Perfil AWS')
    parser.add_argument('--region', type=str, default='eu-west-1', help='Región AWS')
    parser.add_argument('--umbral-hoja', type=float, default=20.0,
                        help='Coste mínimo (US$) para que un servicio tenga hoja propia (default: 20)')
    parser.add_argument('--partner', action='store_true', help='Aplicar descuento de partner')
    parser.add_argument('--descuento', type=float, default=5.0, help='Porcentaje de descuento (default: 5.0)')
    args = parser.parse_args()

    if (args.mes and not args.anio) or (args.anio and not args.mes):
        print("❌ Debes especificar mes Y año, o ninguno")
        sys.exit(1)

    print("=" * 70)
    print("AWS COST REPORT - Una hoja por servicio (estilos + filtros + gráficas)")
    if args.partner:
        print(f"🤝 Modo Partner activado - Descuento: {args.descuento}%")
    print("=" * 70)

    fecha_inicio, fecha_fin = obtener_rango_fechas(args.mes, args.anio)

    session_params = {'region_name': args.region}
    if args.profile:
        session_params['profile_name'] = args.profile
    try:
        session = boto3.Session(**session_params)
        ce = session.client('ce')
        print(f"✅ Conectado a AWS ({args.region})")
    except Exception as e:
        print(f"❌ Error conectando: {e}")
        sys.exit(1)

    costos_base = obtener_costos_base(ce, fecha_inicio, fecha_fin)

    names_con_ec2 = {name for name, servs in costos_base.items()
                     if any(s in servs for s in SERVICIOS_EC2)}
    print(f"   → {len(names_con_ec2)} Names con costos EC2 detectados")
    desglose_ec2 = obtener_desglose_ec2_completo(ce, fecha_inicio, fecha_fin, names_con_ec2)
    ec2_data = normalizar_desglose_ec2(costos_base, desglose_ec2)

    servicios_data = reorganizar_por_servicio(costos_base)
    con_hoja, otros = clasificar_servicios(servicios_data, args.umbral_hoja)

    # Total por Name (para la gráfica Top Names)
    totales_name = sorted(
        ((name, sum(servs.values())) for name, servs in costos_base.items()),
        key=lambda x: x[1], reverse=True)

    print(f"\n📊 {len(con_hoja)} servicios con hoja propia, {len(otros)} agrupados en 'Otros'")

    # Verificación de reconciliación
    total_base = sum(sum(s.values()) for s in costos_base.values())
    total_calc = (sum(sum(c.values()) for c in ec2_data.values())
                  + sum(sum(n.values()) for n in servicios_data.values()))
    print("\n" + "=" * 70)
    print("✅ VERIFICACIÓN:")
    print(f"   Total Cost Explorer (base): ${total_base:,.2f}")
    print(f"   Total calculado (EC2+resto): ${total_calc:,.2f}")
    diff = abs(total_base - total_calc)
    print(f"   {'✅ COINCIDENCIA' if diff < 1 else '⚠️  Diferencia'}: ${diff:,.2f}")
    print("=" * 70)

    crear_excel(ec2_data, con_hoja, otros, totales_name, fecha_inicio, fecha_fin,
                args.output, args.partner, args.descuento)

    print("=" * 70)
    print("✨ Completado exitosamente")
    print("=" * 70)


if __name__ == '__main__':
    main()