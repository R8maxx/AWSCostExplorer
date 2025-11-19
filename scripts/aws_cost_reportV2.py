#!/usr/bin/env python3
"""
Script para extraer costos de AWS por etiqueta Name
- Desglose completo de EC2 sin "Others"
- ServerGroup como información adicional
- AWS Backup por Name automático
"""

import boto3
import pandas as pd
from datetime import datetime
from collections import defaultdict
import argparse
import sys

def obtener_rango_fechas(mes=None, anio=None):
    """Obtiene el rango de fechas para la consulta"""
    if mes and anio:
        fecha_inicio = datetime(anio, mes, 1)
    else:
        ahora = datetime.now()
        fecha_inicio = datetime(ahora.year, ahora.month, 1)
    
    if fecha_inicio.month == 12:
        fecha_fin = datetime(fecha_inicio.year + 1, 1, 1)
    else:
        fecha_fin = datetime(fecha_inicio.year, fecha_inicio.month + 1, 1)
    
    return fecha_inicio.strftime('%Y-%m-%d'), fecha_fin.strftime('%Y-%m-%d')

def obtener_costos_base(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene todos los costos agrupados por servicio y Name"""
    print("📊 Obteniendo costos base por Name y Servicio...")
    
    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'DIMENSION', 'Key': 'SERVICE'},
                {'Type': 'TAG', 'Key': 'Name'}
            ]
        )
        
        costos = defaultdict(lambda: defaultdict(float))
        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                servicio = grupo['Keys'][0]
                name = grupo['Keys'][1].replace('Name$', '') if grupo['Keys'][1] != 'Name$' else 'Sin etiqueta'
                costo = float(grupo['Metrics']['UnblendedCost']['Amount'])
                
                if costo > 0:
                    costos[name][servicio] += costo
        
        return costos
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)

def obtener_servergroup(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene la etiqueta ServerGroup para cada Name"""
    print("🏷️  Obteniendo etiquetas ServerGroup...")
    
    servergroups = {}
    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'TAG', 'Key': 'Name'},
                {'Type': 'TAG', 'Key': 'ServerGroup'}
            ]
        )
        
        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][0] != 'Name$' else ''
                sg = grupo['Keys'][1].replace('ServerGroup$', '') if len(grupo['Keys']) > 1 and grupo['Keys'][1] != 'ServerGroup$' else ''
                
                if name and name != 'Sin etiqueta' and sg:
                    servergroups[name] = sg
        
        return servergroups
    except Exception as e:
        print(f"⚠️  Advertencia ServerGroup: {e}")
        return {}

def obtener_desglose_ec2_completo(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene el desglose COMPLETO de EC2 por Usage Type"""
    print("🔍 Desglosando EC2 en detalle...")
    
    desglose = defaultdict(lambda: defaultdict(float))
    
    # Todos los servicios relacionados con EC2
    servicios_ec2 = [
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    ]
    
    for servicio in servicios_ec2:
        print(f"   → {servicio}")
        try:
            response = cliente_ce.get_cost_and_usage(
                TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
                Granularity='MONTHLY',
                Metrics=['UnblendedCost'],
                Filter={'Dimensions': {'Key': 'SERVICE', 'Values': [servicio]}},
                GroupBy=[
                    {'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'},
                    {'Type': 'TAG', 'Key': 'Name'}
                ]
            )
            
            for periodo in response['ResultsByTime']:
                for grupo in periodo['Groups']:
                    usage_type = grupo['Keys'][0]
                    name = grupo['Keys'][1].replace('Name$', '') if grupo['Keys'][1] != 'Name$' else 'Sin etiqueta'
                    costo = float(grupo['Metrics']['UnblendedCost']['Amount'])
                    
                    if costo > 0:
                        categoria = categorizar_usage_type(usage_type)
                        desglose[name][categoria] += costo
        
        except Exception as e:
            print(f"   ⚠️  {e}")
            continue
    
    return desglose

def categorizar_usage_type(usage_type):
    """Categoriza los Usage Types de EC2 en nombres descriptivos"""
    ut = usage_type.lower()
    
    # Instancias EC2
    if any(x in ut for x in ['boxusage', 'instanceusage', 'hoursusage']):
        # Extraer tipo de instancia si es posible
        if ':' in usage_type:
            tipo = usage_type.split(':')[-1]
            return f'EC2 - Instancia ({tipo})'
        return 'EC2 - Instancias'
    
    # EBS Volumes por tipo
    elif 'volumeusage' in ut:
        if 'gp2' in ut:
            return 'EC2 - EBS Volumes (gp2)'
        elif 'gp3' in ut:
            return 'EC2 - EBS Volumes (gp3)'
        elif 'io1' in ut or 'io2' in ut:
            return 'EC2 - EBS Volumes (io1/io2)'
        elif 'st1' in ut:
            return 'EC2 - EBS Volumes (st1)'
        elif 'sc1' in ut:
            return 'EC2 - EBS Volumes (sc1)'
        else:
            return 'EC2 - EBS Volumes'
    
    # Snapshots
    elif 'snapshot' in ut:
        return 'EC2 - EBS Snapshots'
    
    # IOPS provisionadas
    elif 'piops' in ut or 'volumeiops' in ut:
        return 'EC2 - EBS IOPS'
    
    # Throughput
    elif 'throughput' in ut:
        return 'EC2 - EBS Throughput'
    
    # Network Interfaces
    elif 'networkinterface' in ut or 'createnetworkinterface' in ut:
        return 'EC2 - Network Interfaces (ENI)'
    
    # Elastic IPs
    elif 'elasticip' in ut or 'idleaddress' in ut or 'addressusage' in ut:
        return 'EC2 - Elastic IPs'
    
    # Data Transfer
    elif 'datatransfer' in ut or 'data-transfer' in ut:
        if 'in-bytes' in ut or 'regional-bytes' in ut:
            return 'EC2 - Data Transfer (Regional/In)'
        elif 'out-bytes' in ut or 'bytes' in ut:
            return 'EC2 - Data Transfer (Out)'
        else:
            return 'EC2 - Data Transfer'
    
    # NAT Gateway
    elif 'natgateway' in ut:
        if 'bytes' in ut:
            return 'EC2 - NAT Gateway (Data Processed)'
        else:
            return 'EC2 - NAT Gateway (Hours)'
    
    # Load Balancers
    elif 'loadbalancer' in ut or 'elb:' in ut or 'lcu' in ut:
        if 'application' in ut or 'alb' in ut:
            return 'EC2 - Load Balancer (ALB)'
        elif 'network' in ut or 'nlb' in ut:
            return 'EC2 - Load Balancer (NLB)'
        else:
            return 'EC2 - Load Balancer'
    
    # VPN
    elif 'vpn' in ut:
        return 'EC2 - VPN Connection'
    
    # EBS Optimized
    elif 'ebsoptimized' in ut:
        return 'EC2 - EBS Optimized'
    
    # Spot Instances
    elif 'spot' in ut:
        return 'EC2 - Spot Instances'
    
    # CloudWatch
    elif 'cloudwatch' in ut or 'gmdetailedmonitoring' in ut:
        return 'EC2 - CloudWatch Monitoring'
    
    # Si no se puede categorizar, mostrar el usage type
    else:
        # Limpiar y acortar
        tipo_limpio = usage_type.replace('USE1-', '').replace('EUW1-', '')
        if len(tipo_limpio) > 50:
            return f'EC2 - {tipo_limpio[:50]}...'
        return f'EC2 - {tipo_limpio}'

def obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin):
    """Obtiene costos de AWS Backup por Name (sin necesidad de etiqueta especial)"""
    print("💾 Obteniendo costos de AWS Backup...")
    
    backup_costs = defaultdict(float)
    
    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={'Start': fecha_inicio, 'End': fecha_fin},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            Filter={'Dimensions': {'Key': 'SERVICE', 'Values': ['AWS Backup']}},
            GroupBy=[{'Type': 'TAG', 'Key': 'Name'}]
        )
        
        for periodo in response['ResultsByTime']:
            for grupo in periodo['Groups']:
                name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][0] != 'Name$' else 'Sin etiqueta'
                costo = float(grupo['Metrics']['UnblendedCost']['Amount'])
                
                if costo > 0:
                    backup_costs[name] += costo
        
        return backup_costs
    except Exception as e:
        print(f"⚠️  Advertencia Backup: {e}")
        return {}

def procesar_datos(costos_base, desglose_ec2, backup_costs, servergroups):
    """Procesa y combina todos los datos"""
    print("⚙️  Procesando datos...")
    
    datos_finales = defaultdict(lambda: {'servergroup': '', 'servicios': defaultdict(float)})
    
    # Servicios EC2 que serán reemplazados por el desglose
    servicios_ec2_a_reemplazar = {
        'Amazon Elastic Compute Cloud - Compute',
        'EC2 - Other',
        'Amazon Elastic Block Store'
    }
    
    for name, servicios in costos_base.items():
        # Agregar ServerGroup
        datos_finales[name]['servergroup'] = servergroups.get(name, '')
        
        # Agregar servicios
        for servicio, costo in servicios.items():
            # Si es un servicio EC2 que tenemos desglosado, saltarlo
            if servicio in servicios_ec2_a_reemplazar and name in desglose_ec2:
                continue
            datos_finales[name]['servicios'][servicio] += costo
        
        # Agregar desglose de EC2
        if name in desglose_ec2:
            for categoria, costo in desglose_ec2[name].items():
                datos_finales[name]['servicios'][categoria] += costo
        
        # Agregar AWS Backup
        if name in backup_costs:
            datos_finales[name]['servicios']['AWS Backup'] += backup_costs[name]
    
    return datos_finales

def crear_excel(datos, fecha_inicio, fecha_fin, nombre_archivo):
    """Crea el archivo Excel con los resultados"""
    print("📝 Creando Excel...")
    
    filas = []
    
    # Ordenar por costo total descendente
    datos_ordenados = sorted(
        datos.items(),
        key=lambda x: sum(x[1]['servicios'].values()),
        reverse=True
    )
    
    for name, info in datos_ordenados:
        total = sum(info['servicios'].values())
        servergroup = info['servergroup']
        
        # Fila de total
        filas.append({
            'Name': name,
            'ServerGroup': servergroup,
            'Servicio': '*** TOTAL ***',
            'Costo (US$)': round(total, 2)
        })
        
        # Servicios ordenados por costo
        for servicio, costo in sorted(info['servicios'].items(), key=lambda x: x[1], reverse=True):
            filas.append({
                'Name': '',
                'ServerGroup': '',
                'Servicio': servicio,
                'Costo (US$)': round(costo, 2)
            })
        
        # Línea en blanco
        filas.append({'Name': '', 'ServerGroup': '', 'Servicio': '', 'Costo (US$)': ''})
    
    df = pd.DataFrame(filas)
    
    # Calcular total general
    costo_total = sum(sum(info['servicios'].values()) for info in datos.values())
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detalle de Costos', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Detalle de Costos']
        
        # Ajustar anchos
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 55
        worksheet.column_dimensions['D'].width = 15
        
        # Formato
        from openpyxl.styles import Font, PatternFill
        
        fill_total = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        font_bold = Font(bold=True, size=11)
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[2].value == '*** TOTAL ***':
                for cell in row:
                    cell.fill = fill_total
                    cell.font = font_bold
        
        # Hoja de resumen
        resumen = [
            ['Periodo', f'{fecha_inicio} a {fecha_fin}'],
            [''],
            ['Name', 'ServerGroup', 'Costo Total (US$)']
        ]
        
        for name, info in datos_ordenados:
            total = sum(info['servicios'].values())
            resumen.append([name, info['servergroup'], round(total, 2)])
        
        resumen.append([''])
        resumen.append(['*** TOTAL GENERAL ***', '', round(costo_total, 2)])
        
        df_resumen = pd.DataFrame(resumen)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
        ws_resumen = writer.sheets['Resumen']
        ws_resumen.column_dimensions['A'].width = 40
        ws_resumen.column_dimensions['B'].width = 25
        ws_resumen.column_dimensions['C'].width = 20
        
        # Formato total general
        last_row = len(resumen) + 1
        for cell in ws_resumen[last_row]:
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            cell.font = Font(bold=True, size=12)
        
        # Hoja por ServerGroup
        if any(info['servergroup'] for info in datos.values()):
            sg_totals = defaultdict(float)
            for name, info in datos.items():
                sg = info['servergroup'] if info['servergroup'] else 'Sin ServerGroup'
                sg_totals[sg] += sum(info['servicios'].values())
            
            sg_data = [['ServerGroup', 'Costo Total (US$)']]
            for sg, total in sorted(sg_totals.items(), key=lambda x: x[1], reverse=True):
                sg_data.append([sg, round(total, 2)])
            
            df_sg = pd.DataFrame(sg_data[1:], columns=sg_data[0])
            df_sg.to_excel(writer, sheet_name='Por ServerGroup', index=False)
            
            ws_sg = writer.sheets['Por ServerGroup']
            ws_sg.column_dimensions['A'].width = 35
            ws_sg.column_dimensions['B'].width = 20
            
            for cell in ws_sg[1]:
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
    
    print(f"\n✅ Excel creado: {nombre_archivo}")
    print(f"💰 Costo total: ${costo_total:,.2f} USD")
    print(f"📊 Recursos: {len(datos)}")
    
    return nombre_archivo

def main():
    parser = argparse.ArgumentParser(description='Extrae costos de AWS por Name con desglose EC2 completo')
    parser.add_argument('--mes', type=int, help='Mes (1-12)')
    parser.add_argument('--anio', type=int, help='Año')
    parser.add_argument('--output', type=str, default='aws_costos_detallados.xlsx', help='Archivo de salida')
    parser.add_argument('--profile', type=str, help='Perfil AWS')
    parser.add_argument('--region', type=str, default='us-east-1', help='Región AWS')
    
    args = parser.parse_args()
    
    if (args.mes and not args.anio) or (args.anio and not args.mes):
        print("❌ Debes especificar mes Y año, o ninguno")
        sys.exit(1)
    
    print("=" * 70)
    print("AWS COST REPORT - Desglose Completo por Name")
    print("=" * 70)
    
    # Obtener fechas
    fecha_inicio, fecha_fin = obtener_rango_fechas(args.mes, args.anio)
    
    # Cliente AWS
    session_params = {'region_name': args.region}
    if args.profile:
        session_params['profile_name'] = args.profile
    
    try:
        session = boto3.Session(**session_params)
        ce = session.client('ce')
        print(f"✅ Conectado a AWS ({args.region})")
    except Exception as e:
        print(f"❌ Error conectando: {e}")
        sys.exit(1)
    
    # Obtener datos
    costos_base = obtener_costos_base(ce, fecha_inicio, fecha_fin)
    servergroups = obtener_servergroup(ce, fecha_inicio, fecha_fin)
    desglose_ec2 = obtener_desglose_ec2_completo(ce, fecha_inicio, fecha_fin)
    backup_costs = obtener_costos_backup(ce, fecha_inicio, fecha_fin)
    
    # Procesar
    datos = procesar_datos(costos_base, desglose_ec2, backup_costs, servergroups)
    
    if not datos:
        print("\n⚠️  No se encontraron costos")
        sys.exit(0)
    
    # Crear Excel
    crear_excel(datos, fecha_inicio, fecha_fin, args.output)
    
    print("=" * 70)
    print("✨ Completado exitosamente")
    print("=" * 70)

if __name__ == '__main__':
    main()