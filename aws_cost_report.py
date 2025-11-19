#!/usr/bin/env python3
"""
Script para extraer costos de AWS por etiqueta Name
Incluye todos los servicios (EC2, S3, RDS, etc.) y AWS Backup
Exporta los resultados a Excel
"""

import boto3
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
import argparse
import sys


def obtener_rango_fechas(mes=None, anio=None):
    """
    Obtiene el rango de fechas para la consulta
    Si no se especifica mes/año, usa el mes actual
    """
    if mes and anio:
        fecha_inicio = datetime(anio, mes, 1)
    else:
        ahora = datetime.now()
        fecha_inicio = datetime(ahora.year, ahora.month, 1)

    # Calcular el último día del mes
    if fecha_inicio.month == 12:
        fecha_fin = datetime(fecha_inicio.year + 1, 1, 1)
    else:
        fecha_fin = datetime(fecha_inicio.year, fecha_inicio.month + 1, 1)

    return fecha_inicio.strftime('%Y-%m-%d'), fecha_fin.strftime('%Y-%m-%d')


def obtener_costos_por_servicio(cliente_ce, fecha_inicio, fecha_fin):
    """
    Obtiene los costos agrupados por servicio y etiqueta Name
    """
    print(f"Obteniendo costos desde {fecha_inicio} hasta {fecha_fin}...")

    try:
        response = cliente_ce.get_cost_and_usage(
            TimePeriod={
                'Start': fecha_inicio,
                'End': fecha_fin
            },
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {'Type': 'DIMENSION', 'Key': 'SERVICE'},
                {'Type': 'TAG', 'Key': 'Name'}
            ]
        )

        return response['ResultsByTime']
    except Exception as e:
        print(f"Error al obtener costos: {e}")
        sys.exit(1)


def obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin):
    """
    Obtiene los costos específicos de AWS Backup con sus etiquetas
    """
    print("Obteniendo costos de AWS Backup...")

    costos_backup = {}

    # Mapeo de planes de backup
    planes_backup = {
        'BackupDia': 'avanza_backup_daily',
        'BackupSemana': 'avanza_backup_weekly',
        'BackupMes': 'avanza-backup-monthly'
    }

    try:
        # Obtener costos de AWS Backup filtrados por etiqueta AWSBackup
        for valor_etiqueta, nombre_plan in planes_backup.items():
            response = cliente_ce.get_cost_and_usage(
                TimePeriod={
                    'Start': fecha_inicio,
                    'End': fecha_fin
                },
                Granularity='MONTHLY',
                Metrics=['UnblendedCost'],
                Filter={
                    'And': [
                        {
                            'Dimensions': {
                                'Key': 'SERVICE',
                                'Values': ['AWS Backup']
                            }
                        },
                        {
                            'Tags': {
                                'Key': 'AWSBackup',
                                'Values': [valor_etiqueta]
                            }
                        }
                    ]
                },
                GroupBy=[
                    {'Type': 'TAG', 'Key': 'Name'}
                ]
            )

            for periodo in response['ResultsByTime']:
                for grupo in periodo['Groups']:
                    etiqueta_name = grupo['Keys'][0].replace('Name$', '') if grupo['Keys'][
                                                                                 0] != 'Name$' else 'Sin etiqueta'
                    costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

                    if costo > 0:
                        clave = f"{etiqueta_name}|{nombre_plan}"
                        costos_backup[clave] = costos_backup.get(clave, 0) + costo

        return costos_backup
    except Exception as e:
        print(f"Advertencia: No se pudieron obtener costos de AWS Backup: {e}")
        return {}


def procesar_costos(resultados, costos_backup):
    """
    Procesa los resultados y los organiza por etiqueta Name y servicio
    """
    datos_procesados = defaultdict(lambda: defaultdict(float))

    for periodo in resultados:
        for grupo in periodo['Groups']:
            # grupo['Keys'] contiene [Servicio, Name]
            servicio = grupo['Keys'][0]
            etiqueta_name = grupo['Keys'][1].replace('Name$', '') if len(grupo['Keys']) > 1 else 'Sin etiqueta'

            if etiqueta_name == 'Name$':
                etiqueta_name = 'Sin etiqueta'

            costo = float(grupo['Metrics']['UnblendedCost']['Amount'])

            if costo > 0:
                datos_procesados[etiqueta_name][servicio] += costo

    # Agregar costos de AWS Backup
    for clave, costo in costos_backup.items():
        etiqueta_name, plan_backup = clave.split('|')
        servicio_backup = f"AWS Backup ({plan_backup})"
        datos_procesados[etiqueta_name][servicio_backup] += costo

    return datos_procesados


def crear_excel(datos_procesados, fecha_inicio, fecha_fin, nombre_archivo):
    """
    Crea un archivo Excel con los datos de costos
    """
    # Crear lista de filas para el DataFrame
    filas = []

    for etiqueta_name, servicios in sorted(datos_procesados.items()):
        total_recurso = sum(servicios.values())

        # Fila de resumen por recurso
        fila_resumen = {
            'Etiqueta Name': etiqueta_name,
            'Servicio': 'TOTAL',
            'Costo (US$)': round(total_recurso, 2)
        }
        filas.append(fila_resumen)

        # Filas de detalle por servicio
        for servicio, costo in sorted(servicios.items(), key=lambda x: x[1], reverse=True):
            fila_detalle = {
                'Etiqueta Name': '',
                'Servicio': servicio,
                'Costo (US$)': round(costo, 2)
            }
            filas.append(fila_detalle)

        # Fila en blanco para separar
        filas.append({'Etiqueta Name': '', 'Servicio': '', 'Costo (US$)': ''})

    # Crear DataFrame
    df = pd.DataFrame(filas)

    # Calcular costo total general
    costo_total = sum(sum(servicios.values()) for servicios in datos_procesados.values())

    # Crear archivo Excel con formato
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Costos por Etiqueta', index=False)

        # Obtener el workbook y worksheet para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets['Costos por Etiqueta']

        # Ajustar anchos de columna
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 15

        # Aplicar formato a las celdas de totales
        from openpyxl.styles import Font, PatternFill, Alignment

        fill_total = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font_bold = Font(bold=True)

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            if row[1].value == 'TOTAL':
                for cell in row:
                    cell.fill = fill_total
                    cell.font = font_bold

        # Agregar hoja de resumen
        resumen = []
        resumen.append(['Periodo', f"{fecha_inicio} a {fecha_fin}"])
        resumen.append([''])
        resumen.append(['Resumen por Etiqueta Name', 'Costo Total (US$)'])

        for etiqueta_name, servicios in sorted(datos_procesados.items(), key=lambda x: sum(x[1].values()),
                                               reverse=True):
            total = sum(servicios.values())
            resumen.append([etiqueta_name, round(total, 2)])

        resumen.append([''])
        resumen.append(['COSTO TOTAL GENERAL', round(costo_total, 2)])

        df_resumen = pd.DataFrame(resumen)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)

        worksheet_resumen = writer.sheets['Resumen']
        worksheet_resumen.column_dimensions['A'].width = 30
        worksheet_resumen.column_dimensions['B'].width = 20

        # Formato para el total general
        last_row = len(resumen) + 1
        for cell in worksheet_resumen[last_row]:
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            cell.font = Font(bold=True, size=12)

    print(f"\n✓ Archivo Excel creado: {nombre_archivo}")
    print(f"✓ Costo total del periodo: ${costo_total:,.2f} US$")
    return nombre_archivo


def main():
    parser = argparse.ArgumentParser(
        description='Extrae costos de AWS por etiqueta Name y exporta a Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  %(prog)s                          # Costos del mes actual
  %(prog)s --mes 10 --anio 2024     # Costos de octubre 2024
  %(prog)s --mes 1 --anio 2025      # Costos de enero 2025
  %(prog)s --output costos_aws.xlsx # Especificar nombre de archivo
        """
    )

    parser.add_argument('--mes', type=int, help='Mes (1-12)')
    parser.add_argument('--anio', type=int, help='Año (ej: 2024)')
    parser.add_argument('--output', type=str, default='aws_costos_por_etiqueta.xlsx',
                        help='Nombre del archivo Excel de salida')
    parser.add_argument('--profile', type=str, help='Perfil de AWS CLI a usar')
    parser.add_argument('--region', type=str, default='us-east-1', help='Región de AWS')

    args = parser.parse_args()

    # Validar mes si se proporciona
    if args.mes and (args.mes < 1 or args.mes > 12):
        print("Error: El mes debe estar entre 1 y 12")
        sys.exit(1)

    # Validar que si se da mes, se dé año y viceversa
    if (args.mes and not args.anio) or (args.anio and not args.mes):
        print("Error: Debe especificar tanto --mes como --anio, o ninguno")
        sys.exit(1)

    print("=" * 60)
    print("AWS Cost Report - Costos por Etiqueta Name")
    print("=" * 60)

    # Obtener rango de fechas
    fecha_inicio, fecha_fin = obtener_rango_fechas(args.mes, args.anio)

    # Crear cliente de Cost Explorer
    try:
        session_params = {'region_name': args.region}
        if args.profile:
            session_params['profile_name'] = args.profile

        session = boto3.Session(**session_params)
        cliente_ce = session.client('ce')
        print(f"✓ Conectado a AWS (región: {args.region})")
    except Exception as e:
        print(f"Error al conectar con AWS: {e}")
        print("\nAsegúrese de tener configuradas sus credenciales de AWS")
        print("Puede usar: aws configure")
        sys.exit(1)

    # Obtener datos de costos
    resultados = obtener_costos_por_servicio(cliente_ce, fecha_inicio, fecha_fin)
    costos_backup = obtener_costos_backup(cliente_ce, fecha_inicio, fecha_fin)

    # Procesar datos
    print("Procesando datos...")
    datos_procesados = procesar_costos(resultados, costos_backup)

    if not datos_procesados:
        print("\n⚠ No se encontraron costos en el periodo especificado")
        sys.exit(0)

    # Crear archivo Excel
    archivo_salida = crear_excel(datos_procesados, fecha_inicio, fecha_fin, args.output)

    print("\n" + "=" * 60)
    print(f"Recursos encontrados: {len(datos_procesados)}")
    print("=" * 60)


if __name__ == '__main__':
    main()
